"""Strict CSV/XLSX import with cell-level provenance and saved mappings."""

from __future__ import annotations

import csv
import io
import json
import math
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, BinaryIO

import pandas as pd

from .provenance import sha256_bytes
from .schema import (
    IMPORT_MODES,
    OPTIONAL_PROTOCOL_COLUMNS,
    REQUIRED_COLUMNS,
    RawCell,
    ValidationIssue,
)


_MEASUREMENT_COLUMNS = (
    "settlement",
    "indicator_1",
    "indicator_2",
    "indicator_3",
    "indicator_4",
)

_MAX_XLSX_INPUT_BYTES = 64 * 1024 * 1024
_MAX_XLSX_ENTRIES = 5_000
_MAX_XLSX_ENTRY_BYTES = 64 * 1024 * 1024
_MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES = 256 * 1024 * 1024
_MAX_XLSX_COMPRESSION_RATIO = 500.0
_MAX_XLSX_SHEETS = 64
_MAX_XLSX_ROWS_PER_SHEET = 200_000
_MAX_XLSX_COLUMNS_PER_SHEET = 512
_REQUIRED_XLSX_ENTRIES = {"[content_types].xml", "xl/workbook.xml"}


class _XlsxGuardError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "test_id": (
        "test_id",
        "test id",
        "испытание",
        "номер испытания",
        "№ испытания",
        "опыт",
        "номер опыта",
    ),
    "stage": ("stage", "ступень", "номер ступени", "№ ступени", "шаг", "этап"),
    "load": (
        "load",
        "load kn",
        "load n",
        "load mn",
        "load kgf",
        "load tf",
        "load pa",
        "load kpa",
        "load mpa",
        "force",
        "force n",
        "force kn",
        "нагрузка",
        "нагрузка кн",
        "нагрузка н",
        "нагрузка мн",
        "нагрузка кгс",
        "нагрузка тс",
        "сила",
        "усилие",
        "f",
        "f кн",
        "давление",
        "давление па",
        "давление кпа",
        "давление мпа",
        "pressure",
        "pressure pa",
        "pressure kpa",
        "pressure mpa",
        "p",
        "p pa",
        "p кпа",
        "p kpa",
        "p mpa",
    ),
    "settlement": ("settlement", "осадка", "осадка мм", "s", "s мм"),
    "indicator_1": (
        "indicator_1",
        "indicator 1",
        "индикатор 1",
        "ич 1",
        "ич-1",
        "показание индикатора",
        "индикатор",
    ),
    "indicator_2": ("indicator_2", "indicator 2", "индикатор 2", "ич 2", "ич-2"),
    "indicator_3": ("indicator_3", "indicator 3", "индикатор 3", "ич 3", "ич-3"),
    "indicator_4": ("indicator_4", "indicator 4", "индикатор 4", "ич 4", "ич-4"),
    "indicator_1_turn_number": (
        "indicator_1_turn_number",
        "indicator 1 turn number",
        "оборот индикатора 1",
        "номер оборота индикатора 1",
    ),
    "indicator_2_turn_number": (
        "indicator_2_turn_number",
        "indicator 2 turn number",
        "оборот индикатора 2",
        "номер оборота индикатора 2",
    ),
    "indicator_3_turn_number": (
        "indicator_3_turn_number",
        "indicator 3 turn number",
        "оборот индикатора 3",
        "номер оборота индикатора 3",
    ),
    "indicator_4_turn_number": (
        "indicator_4_turn_number",
        "indicator 4 turn number",
        "оборот индикатора 4",
        "номер оборота индикатора 4",
    ),
    "reference_indicator": (
        "reference_indicator",
        "reference indicator",
        "реперный индикатор",
        "репер",
    ),
    "horizontal_indicator": (
        "horizontal_indicator",
        "horizontal indicator",
        "горизонтальный индикатор",
        "горизонтальное перемещение",
    ),
    "branch": ("branch", "ветвь", "режим нагружения"),
    "timestamp": ("timestamp", "time", "время", "дата и время"),
    "status": ("status", "статус", "состояние"),
    "comment": ("comment", "комментарий", "примечание"),
    "group": ("group", "группа", "серия"),
    "pair_id": ("pair_id", "pair id", "пара", "номер пары"),
}

_FAILURE_WORDS = ("ушла", "ушел", "ушёл", "разруш", "срыв", "провал", "failure", "failed")
_LEGACY_TITLE_PATTERNS = (
    re.compile(r"^\s*(\d+)\s*(?:[-–—.]\s*)?(?:испыт(?:ание|ания)?|опыт)\b", re.I),
    re.compile(r"^\s*(?:испыт(?:ание)?|опыт)\s*№?\s*(\d+)\b", re.I),
)


@dataclass(slots=True)
class ProtocolImportResult:
    frame: pd.DataFrame
    info: dict[str, Any]
    issues: list[ValidationIssue]
    raw_cells: pd.DataFrame

    @property
    def blocking_issues(self) -> list[ValidationIssue]:
        return [item for item in self.issues if bool(item.blocks_processing)]


def _read_bytes(source: str | Path | bytes | bytearray | BinaryIO) -> bytes:
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    if hasattr(source, "getvalue"):
        data = source.getvalue()
    else:
        position = source.tell() if hasattr(source, "tell") else None
        data = source.read()
        if position is not None and hasattr(source, "seek"):
            source.seek(position)
    return data.encode("utf-8") if isinstance(data, str) else bytes(data)


def _decode(data: bytes, encoding: str | None) -> tuple[str, str]:
    if encoding:
        candidates = [encoding]
    elif data.startswith((b"\xff\xfe", b"\xfe\xff")):
        candidates = ["utf-16", "utf-8-sig", "cp1251"]
    else:
        candidates = ["utf-8-sig", "utf-8", "cp1251"]
    last_error: UnicodeDecodeError | None = None
    for candidate in candidates:
        try:
            return data.decode(candidate), candidate
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError("Не удалось определить кодировку CSV (проверены UTF-8 и CP1251).") from last_error


def _dialect(text: str, delimiter: str | None) -> str:
    if delimiter:
        return delimiter
    sample = "\n".join(text.splitlines()[:20])
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        counts = {candidate: sample.count(candidate) for candidate in (";", "\t", ",", "|")}
        return max(counts, key=counts.get)


def _first_measurement_column(columns: list[str] | pd.Index) -> str | None:
    return next((name for name in _MEASUREMENT_COLUMNS if name in columns), None)


def _attach_source_columns(
    frame: pd.DataFrame,
    raw_frame: pd.DataFrame,
    *,
    sheet_name: str,
    first_source_row: int,
) -> pd.DataFrame:
    result = frame.copy(deep=True)
    result["sheet_name"] = sheet_name
    result["source_row"] = range(first_source_row, first_source_row + len(result))
    result["sequence_index"] = range(len(result))
    source_columns = {str(name): str(name) for name in frame.columns}
    result["source_columns"] = json.dumps(source_columns, ensure_ascii=False, sort_keys=True)
    if "stage" in result:
        result["raw_stage"] = raw_frame["stage"].where(raw_frame["stage"].ne(""), None)
        result["parsed_stage"] = result["stage"]
    if "load" in result:
        result["raw_load"] = raw_frame["load"].where(raw_frame["load"].ne(""), None)
        result["parsed_load"] = pd.to_numeric(result["load"], errors="coerce")
    indicator = _first_measurement_column(list(result.columns))
    if indicator:
        result["raw_indicator"] = raw_frame[indicator].where(raw_frame[indicator].ne(""), None)
        result["parsed_indicator"] = pd.to_numeric(result[indicator], errors="coerce")
    else:
        result["raw_indicator"] = None
        result["parsed_indicator"] = math.nan
    result["source_load_unit"] = None
    result["load_unit"] = None
    return result


def read_protocol_csv(
    source: str | Path | bytes | bytearray | BinaryIO,
    *,
    encoding: str | None = None,
    delimiter: str | None = None,
    decimal: str | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    data = _read_bytes(source)
    text, actual_encoding = _decode(data, encoding)
    actual_delimiter = _dialect(text, delimiter)
    if decimal is None:
        decimal = "," if actual_delimiter == ";" else "."
    raw_frame = pd.read_csv(
        io.StringIO(text),
        sep=actual_delimiter,
        dtype=str,
        keep_default_na=False,
    )
    frame = pd.read_csv(
        io.StringIO(text),
        sep=actual_delimiter,
        decimal=decimal,
        dtype={"test_id": "string", "branch": "string", "status": "string", "comment": "string"},
    )
    frame.columns = [str(name).strip() for name in frame.columns]
    raw_frame.columns = [str(name).strip() for name in raw_frame.columns]
    frame = _attach_source_columns(frame, raw_frame, sheet_name="CSV", first_source_row=2)
    info = {
        "format": "csv",
        "encoding": actual_encoding,
        "delimiter": actual_delimiter,
        "decimal": decimal,
        "rows": len(frame),
        "columns": list(frame.columns),
        "source_columns": list(raw_frame.columns),
        "input_file_sha256": sha256_bytes(data),
        "import_mode": "strict",
    }
    return frame, info


def read_metadata_json(source: str | Path | bytes | bytearray | BinaryIO) -> dict[str, Any]:
    data = _read_bytes(source)
    text, _ = _decode(data, None)

    def reject_duplicate_keys(pairs):
        result = {}
        for key, value in pairs:
            if key in result:
                raise ValueError(f"Metadata JSON содержит повторяющийся ключ {key!r}.")
            result[key] = value
        return result

    parsed = json.loads(text, object_pairs_hook=reject_duplicate_keys)
    if not isinstance(parsed, dict):
        raise ValueError("Корень metadata JSON должен быть объектом.")
    return parsed


_UNIT_CANONICAL = {
    "n": "N",
    "н": "N",
    "kn": "kN",
    "кн": "kN",
    "mn": "MN",
    "мн": "MN",
    "kgf": "kgf",
    "кгс": "kgf",
    "tf": "tf",
    "тс": "tf",
    "pa": "Pa",
    "па": "Pa",
    "kpa": "kPa",
    "кпа": "kPa",
    "mpa": "MPa",
    "мпа": "MPa",
}
_FORCE_UNITS = {"N", "kN", "MN", "kgf", "tf"}
_PRESSURE_UNITS = {"Pa", "kPa", "MPa"}


def validate_import_metadata_consistency(
    frame: pd.DataFrame,
    metadata: dict[str, Any],
    import_info: dict[str, Any] | None = None,
    *,
    strict: bool = True,
) -> list[ValidationIssue]:
    """Reject a conflict between an explicit XLSX header unit and metadata."""

    issues: list[ValidationIssue] = []
    metadata_tests = metadata.get("tests")
    if isinstance(metadata_tests, dict) and metadata_tests and "test_id" in frame:
        protocol_ids = set(frame["test_id"].dropna().astype(str))
        metadata_ids = {str(value) for value in metadata_tests}
        missing_ids = sorted(protocol_ids - metadata_ids)
        extra_ids = sorted(metadata_ids - protocol_ids)
        if missing_ids:
            issues.append(
                ValidationIssue(
                    "error" if strict else "warning",
                    "protocol_tests_missing_in_metadata",
                    "В metadata.tests отсутствуют испытания протокола: " + ", ".join(missing_ids),
                    raw_value=missing_ids,
                    suggested_action="Загрузите metadata этого проекта или добавьте паспорта испытаний.",
                )
            )
        if extra_ids:
            issues.append(
                ValidationIssue(
                    "warning",
                    "metadata_tests_not_in_protocol",
                    "В metadata.tests есть отсутствующие в протоколе испытания: " + ", ".join(extra_ids),
                    raw_value=extra_ids,
                    suggested_action="Проверьте, что выбран правильный протокол и полный набор листов.",
                )
            )
    unit_column = "source_load_unit" if "source_load_unit" in frame else "load_unit"
    if unit_column not in frame:
        return issues
    sheets = (import_info or {}).get("sheets") or []

    grouped = frame.groupby(frame["test_id"].astype(str), sort=False) if "test_id" in frame else [(None, frame)]
    for test_id, part in grouped:
        header_units = {
            _UNIT_CANONICAL.get(str(value).strip().casefold(), str(value).strip())
            for value in part[unit_column].dropna().unique()
            if str(value).strip()
        }
        if not header_units:
            continue
        part_sheet = str(part["sheet_name"].iloc[0]) if "sheet_name" in part and len(part) else None
        sheet_info = next(
            (item for item in sheets if item.get("sheet_name") == part_sheet),
            sheets[0] if sheets else {},
        )
        row = sheet_info.get("header_row")
        column = (sheet_info.get("mapping") or {}).get("load")
        if len(header_units) > 1:
            issues.append(
                ValidationIssue(
                    "error",
                    "multiple_load_units_in_test",
                    f"В испытании {test_id} обнаружены разные единицы нагрузки: "
                    + ", ".join(sorted(header_units)),
                    test_id=str(test_id) if test_id is not None else None,
                    sheet=part_sheet,
                    row=row,
                    column=column,
                    suggested_action="Используйте одну явно заданную единицу внутри испытания.",
                )
            )
            continue
        effective = dict(metadata)
        if isinstance(metadata_tests, dict) and str(test_id) in metadata_tests:
            override = metadata_tests[str(test_id)]
            if isinstance(override, dict):
                effective.update(override)
        header_unit = next(iter(header_units))
        metadata_raw = effective.get("load_unit")
        metadata_unit = _UNIT_CANONICAL.get(
            str(metadata_raw).strip().casefold(), str(metadata_raw).strip()
        )
        if metadata_unit != header_unit:
            issues.append(
                ValidationIssue(
                    "error",
                    "load_unit_conflict",
                    f"Заголовок Excel задаёт {header_unit}, а effective metadata.load_unit для {test_id} — {metadata_raw!r}.",
                    test_id=str(test_id) if test_id is not None else None,
                    sheet=part_sheet,
                    row=row,
                    column=column,
                    raw_value=header_unit,
                    suggested_action="Согласуйте единицу заголовка с глобальной/test-specific metadata.",
                )
            )
        metadata_kind = str(effective.get("load_kind", "force")).strip().casefold()
        header_kind = (
            "force"
            if header_unit in _FORCE_UNITS
            else "pressure" if header_unit in _PRESSURE_UNITS else None
        )
        normalized_kind = (
            "pressure" if metadata_kind in {"pressure", "p", "давление"} else "force"
        )
        if header_kind and header_kind != normalized_kind:
            issues.append(
                ValidationIssue(
                    "error",
                    "load_kind_conflict",
                    f"Единица {header_unit} соответствует {header_kind}, а effective load_kind для {test_id} — {effective.get('load_kind')!r}.",
                    test_id=str(test_id) if test_id is not None else None,
                    sheet=part_sheet,
                    row=row,
                    column=column,
                    raw_value=header_unit,
                    suggested_action="Исправьте test-specific load_kind или единицу исходного столбца.",
                )
            )
    return issues


def _normalise_header(value: Any) -> str:
    text = str(value or "").strip().casefold().replace("ё", "е")
    text = re.sub(r"[\[\](){}]", " ", text)
    text = re.sub(r"[_/\\,;:.№-]+", " ", text)
    return " ".join(text.split())


_NORMALISED_ALIASES = {
    canonical: {_normalise_header(alias) for alias in aliases} | {_normalise_header(canonical)}
    for canonical, aliases in _HEADER_ALIASES.items()
}


def _canonical_header(value: Any) -> str | None:
    normalized = _normalise_header(value)
    if not normalized:
        return None
    for canonical, aliases in _NORMALISED_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


def parse_decimal(value: Any, *, family: str | None = None) -> float | None:
    """Parse one finite decimal value without changing its physical unit.

    Both decimal comma and decimal point are accepted.  ``family`` applies
    the same cell-safety rules as the strict Excel importer: unit suffixes in
    primary cells are rejected because their conversion belongs to explicit
    metadata, not to text parsing.
    """

    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    match = re.fullmatch(
        r"([+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)"
        r"([a-zа-я]+)?",
        text,
        flags=re.I,
    )
    if not match:
        return None
    unit = (match.group(2) or "").casefold()
    load_units = {"кн", "kn", "н", "n", "мн", "mn", "кгс", "kgf", "тс", "tf", "кпа", "kpa", "па", "pa", "мпа", "mpa"}
    measurement_units = {"мм", "mm", "см", "cm", "м", "m"}
    if family in {"load", "measurement", "stage"} and unit:
        # TASK 01 keeps units in headers/metadata. Silently stripping a cell
        # suffix would make 1000 N become 1000 kN or 0.1 cm become 0.1 mm.
        return None
    if family == "load" and unit and unit not in load_units:
        return None
    if family == "measurement" and unit and unit not in measurement_units:
        return None
    if family == "stage" and unit:
        return None
    if family is None and unit and unit not in load_units | measurement_units:
        return None
    number = float(match.group(1))
    return number if math.isfinite(number) else None


# Private compatibility name retained for the established import pipeline.
_parse_number = parse_decimal


def _is_failure(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = " ".join(value.strip().casefold().replace("ё", "е").split())
    if re.search(r"\b(?:не|not|без)\s+(?:ушла|ушел|разруш\w*|срыв\w*|провал\w*|fail\w*)", text):
        return False
    if re.search(r"(?:разруш\w*|срыв\w*|провал\w*)\s+(?:нет|не было|отсутств\w*)", text):
        return False
    return any(word in text for word in _FAILURE_WORDS)


def _infer_load_unit(header: Any) -> str | None:
    text = _normalise_header(header)
    for token, unit in (
        ("мпа", "MPa"),
        ("mpa", "MPa"),
        ("кпа", "kPa"),
        ("kpa", "kPa"),
        ("кн", "kN"),
        ("kn", "kN"),
        ("мн", "MN"),
        ("mn", "MN"),
        ("кгс", "kgf"),
        ("kgf", "kgf"),
        ("тс", "tf"),
        ("tf", "tf"),
        ("па", "Pa"),
        ("pa", "Pa"),
        ("н", "N"),
        ("n", "N"),
    ):
        if token in text.split():
            return unit
    return None


def _validate_xlsx_archive(data: bytes) -> None:
    if len(data) > _MAX_XLSX_INPUT_BYTES:
        raise _XlsxGuardError(
            "xlsx_resource_limit",
            f"Размер XLSX превышает {_MAX_XLSX_INPUT_BYTES // (1024 * 1024)} MiB.",
        )
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            infos = archive.infolist()
            if len(infos) > _MAX_XLSX_ENTRIES:
                raise _XlsxGuardError(
                    "xlsx_resource_limit",
                    f"В XLSX слишком много ZIP entries: {len(infos)}.",
                )
            normalized_names = [item.filename.replace("\\", "/").casefold() for item in infos]
            if len(normalized_names) != len(set(normalized_names)):
                raise _XlsxGuardError(
                    "invalid_xlsx_container",
                    "XLSX содержит повторяющиеся ZIP entries.",
                )
            missing = sorted(_REQUIRED_XLSX_ENTRIES - set(normalized_names))
            if missing:
                raise _XlsxGuardError(
                    "invalid_xlsx_container",
                    "В контейнере отсутствуют обязательные части XLSX: " + ", ".join(missing),
                )
            total_uncompressed = 0
            allowed_compression = {zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED}
            for item in infos:
                if item.flag_bits & 0x1:
                    raise _XlsxGuardError(
                        "invalid_xlsx_container",
                        f"Зашифрованный ZIP entry не поддерживается: {item.filename!r}.",
                    )
                if item.compress_type not in allowed_compression:
                    raise _XlsxGuardError(
                        "invalid_xlsx_container",
                        f"Неподдерживаемый метод ZIP-сжатия: {item.filename!r}.",
                    )
                if item.file_size > _MAX_XLSX_ENTRY_BYTES:
                    raise _XlsxGuardError(
                        "xlsx_resource_limit",
                        f"ZIP entry слишком велик после распаковки: {item.filename!r}.",
                    )
                total_uncompressed += item.file_size
                if total_uncompressed > _MAX_XLSX_TOTAL_UNCOMPRESSED_BYTES:
                    raise _XlsxGuardError(
                        "xlsx_resource_limit",
                        "Суммарный распакованный объём XLSX превышает безопасный лимит.",
                    )
                if item.file_size >= 1024 * 1024:
                    ratio = item.file_size / max(item.compress_size, 1)
                    if ratio > _MAX_XLSX_COMPRESSION_RATIO:
                        raise _XlsxGuardError(
                            "xlsx_resource_limit",
                            f"Подозрительный коэффициент ZIP-сжатия у {item.filename!r}: {ratio:.1f}.",
                        )
    except zipfile.BadZipFile as exc:
        raise _XlsxGuardError(
            "invalid_xlsx_container", "Файл не является корректным ZIP/XLSX-контейнером."
        ) from exc


def _validate_workbook_limits(workbook) -> None:
    if len(workbook.sheetnames) > _MAX_XLSX_SHEETS:
        raise _XlsxGuardError(
            "xlsx_resource_limit",
            f"В книге слишком много листов: {len(workbook.sheetnames)}.",
        )
    for worksheet in workbook.worksheets:
        max_row = int(worksheet.max_row or 0)
        max_column = int(worksheet.max_column or 0)
        if max_row > _MAX_XLSX_ROWS_PER_SHEET or max_column > _MAX_XLSX_COLUMNS_PER_SHEET:
            raise _XlsxGuardError(
                "xlsx_resource_limit",
                f"Заявленный размер листа {worksheet.title!r} ({max_row}×{max_column}) превышает безопасный лимит.",
            )


def _xlsx_exception_code(exc: Exception) -> str:
    current: BaseException | None = exc
    seen: set[int] = set()
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        module = current.__class__.__module__.casefold()
        name = current.__class__.__name__.casefold()
        if module.startswith("defusedxml") or "forbidden" in name:
            return "unsafe_xml"
        current = current.__cause__ or current.__context__
    return "invalid_xlsx_workbook"


def _excel_modules():
    try:
        from openpyxl.xml import DEFUSEDXML, LXML
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Для импорта XLSX установите зависимость openpyxl.") from exc
    if not (DEFUSEDXML or LXML):
        raise RuntimeError(
            "Безопасный XML parser для XLSX недоступен; установите defusedxml>=0.7.1."
        )
    return load_workbook, get_column_letter


def _row_headers(ws, row_number: int) -> list[tuple[int, str, Any]]:
    load_workbook, get_column_letter = _excel_modules()
    del load_workbook
    return [
        (cell.column, get_column_letter(cell.column), cell.value)
        for cell in ws[row_number]
        if cell.value is not None and str(cell.value).strip()
    ]


def _discover_header(
    ws, explicit_row: int | None = None, *, allow_unrecognized: bool = False
) -> dict[str, Any] | None:
    rows = [explicit_row] if explicit_row else range(1, min(int(ws.max_row or 0), 50) + 1)
    best: dict[str, Any] | None = None
    for row_number in rows:
        headers = _row_headers(ws, row_number)
        recognized: dict[str, list[tuple[int, str, Any]]] = {}
        for index, letter, value in headers:
            canonical = _canonical_header(value)
            if canonical:
                recognized.setdefault(canonical, []).append((index, letter, value))
        score = sum(10 for name in REQUIRED_COLUMNS if name in recognized)
        score += sum(5 for name in _MEASUREMENT_COLUMNS if name in recognized)
        score += len(recognized)
        candidate = {
            "header_row": row_number,
            "headers": headers,
            "recognized": recognized,
            "score": score,
        }
        if best is None or candidate["score"] > best["score"]:
            best = candidate
    if best is None or not best["headers"]:
        return None
    if best["score"] < 20 and not (allow_unrecognized or explicit_row is not None):
        return None
    return best


def _looks_like_tabular_sheet(ws) -> bool:
    populated_rows = 0
    for row in ws.iter_rows(min_row=1, max_row=min(int(ws.max_row or 0), 50)):
        if sum(cell.value is not None and str(cell.value).strip() != "" for cell in row) >= 2:
            populated_rows += 1
            if populated_rows >= 2:
                return True
    return False


def inspect_excel_schema(
    source: str | Path | bytes | bytearray | BinaryIO,
    *,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> dict[str, Any]:
    data = _read_bytes(source)
    _validate_xlsx_archive(data)
    load_workbook, _ = _excel_modules()
    workbook = None
    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=False, keep_links=False
        )
        _validate_workbook_limits(workbook)
    except _XlsxGuardError:
        if workbook is not None:
            workbook.close()
        raise
    except Exception as exc:
        if workbook is not None:
            workbook.close()
        raise _XlsxGuardError(
            _xlsx_exception_code(exc), f"Не удалось безопасно разобрать структуру XLSX: {exc}"
        ) from exc
    try:
        selected = [sheet_name] if sheet_name else workbook.sheetnames
        sheets = []
        for name in selected:
            if name not in workbook.sheetnames:
                raise ValueError(f"Лист {name!r} отсутствует в книге.")
            discovery = _discover_header(
                workbook[name], header_row, allow_unrecognized=True
            )
            if discovery is None:
                sheets.append({"sheet_name": name, "header_row": None, "headers": [], "suggested_mapping": {}})
                continue
            suggested = {
                canonical: items[0][1]
                for canonical, items in discovery["recognized"].items()
                if len(items) == 1
            }
            sheets.append(
                {
                    "sheet_name": name,
                    "header_row": discovery["header_row"],
                    "headers": [
                        {"column": letter, "value": value} for _, letter, value in discovery["headers"]
                    ],
                    "suggested_mapping": suggested,
                }
            )
        return {"input_file_sha256": sha256_bytes(data), "sheets": sheets}
    finally:
        workbook.close()


def _resolve_mapping(
    discovery: dict[str, Any],
    *,
    mode: str,
    supplied: dict[str, str | int] | None,
    sheet_name: str,
) -> tuple[dict[str, int], dict[str, str], list[ValidationIssue]]:
    issues: list[ValidationIssue] = []
    headers = discovery["headers"]
    by_letter = {letter.casefold(): (index, value) for index, letter, value in headers}
    by_text = {_normalise_header(value): (index, letter, value) for index, letter, value in headers}
    mapping: dict[str, int] = {}
    source_letters: dict[str, str] = {}

    if mode == "interactive":
        for canonical, selector in (supplied or {}).items():
            if canonical not in (*REQUIRED_COLUMNS, *OPTIONAL_PROTOCOL_COLUMNS, "pair_id"):
                issues.append(
                    ValidationIssue(
                        "warning",
                        "unknown_mapping_target",
                        f"Неизвестное целевое поле mapping: {canonical}.",
                        sheet=sheet_name,
                        suggested_action="Удалите поле из mapping JSON.",
                    )
                )
                continue
            resolved: tuple[int, str, Any] | None = None
            if isinstance(selector, int):
                match = next((item for item in headers if item[0] == selector), None)
                resolved = match
            else:
                key = str(selector).strip().casefold()
                if key in by_letter:
                    index, value = by_letter[key]
                    resolved = (index, str(selector).upper(), value)
                else:
                    candidate = by_text.get(_normalise_header(selector))
                    if candidate:
                        index, letter, value = candidate
                        resolved = (index, letter, value)
            if resolved is None:
                issues.append(
                    ValidationIssue(
                        "error",
                        "mapping_source_not_found",
                        f"Столбец {selector!r} для поля {canonical} не найден.",
                        sheet=sheet_name,
                        column=str(selector),
                        suggested_action="Выберите существующий столбец из предпросмотра.",
                    )
                )
                continue
            mapping[canonical] = resolved[0]
            source_letters[canonical] = resolved[1]
    else:
        for canonical, items in discovery["recognized"].items():
            if len(items) > 1:
                issues.append(
                    ValidationIssue(
                        "error" if mode == "strict" else "warning",
                        "ambiguous_header",
                        f"Несколько столбцов распознаны как {canonical}.",
                        sheet=sheet_name,
                        row=discovery["header_row"],
                        column=", ".join(item[1] for item in items),
                        suggested_action="Используйте interactive mapping.",
                    )
                )
                continue
            index, letter, _ = items[0]
            mapping[canonical] = index
            source_letters[canonical] = letter

        recognized_indices = set(mapping.values())
        unknown = [(index, letter, value) for index, letter, value in headers if index not in recognized_indices]
        for _, letter, value in unknown:
            issues.append(
                ValidationIssue(
                    "error" if mode == "strict" else "warning",
                    "unknown_header",
                    f"Неизвестный заголовок {value!r}.",
                    sheet=sheet_name,
                    row=discovery["header_row"],
                    column=letter,
                    raw_value=value,
                    suggested_action="Переименуйте столбец или используйте interactive mapping.",
                )
            )

    for required in REQUIRED_COLUMNS:
        if required not in mapping:
            issues.append(
                ValidationIssue(
                    "error",
                    "missing_required_mapping",
                    f"Не сопоставлено обязательное поле {required}.",
                    sheet=sheet_name,
                    row=discovery["header_row"],
                    suggested_action="Укажите соответствующий столбец в mapping.",
                )
            )
    if not any(name in mapping for name in _MEASUREMENT_COLUMNS):
        issues.append(
            ValidationIssue(
                "error",
                "missing_measurement_mapping",
                "Не сопоставлена осадка или хотя бы один индикатор.",
                sheet=sheet_name,
                row=discovery["header_row"],
                suggested_action="Сопоставьте settlement или indicator_1..4.",
            )
        )
    duplicate_sources = {index for index in mapping.values() if list(mapping.values()).count(index) > 1}
    for index in duplicate_sources:
        fields = [name for name, value in mapping.items() if value == index]
        issues.append(
            ValidationIssue(
                "error",
                "duplicate_mapping_source",
                "Один исходный столбец назначен нескольким полям: " + ", ".join(fields),
                sheet=sheet_name,
                row=discovery["header_row"],
                suggested_action="Назначьте каждому полю отдельный столбец.",
            )
        )
    return mapping, source_letters, issues


def _row_value(ws, row_number: int, column_number: int) -> Any:
    return ws.cell(row=row_number, column=column_number).value


def _numeric_field(name: str) -> bool:
    return name in {
        "load",
        "settlement",
        "indicator_1",
        "indicator_2",
        "indicator_3",
        "indicator_4",
        "indicator_1_turn_number",
        "indicator_2_turn_number",
        "indicator_3_turn_number",
        "indicator_4_turn_number",
        "reference_indicator",
        "horizontal_indicator",
    }


def _parse_long_table_sheet(
    ws,
    discovery: dict[str, Any],
    mapping: dict[str, int],
    source_letters: dict[str, str],
    *,
    sequence_start: int,
    raw_ws=None,
) -> tuple[list[dict[str, Any]], list[RawCell], list[ValidationIssue]]:
    records: list[dict[str, Any]] = []
    cells: list[RawCell] = []
    issues: list[ValidationIssue] = []
    raw_ws = raw_ws or ws
    header_values = {
        name: _row_value(raw_ws, discovery["header_row"], column)
        for name, column in mapping.items()
    }
    load_unit = _infer_load_unit(header_values.get("load"))
    for row_number in range(discovery["header_row"] + 1, int(ws.max_row or 0) + 1):
        raw = {name: _row_value(raw_ws, row_number, column) for name, column in mapping.items()}
        data_values = {name: _row_value(ws, row_number, column) for name, column in mapping.items()}
        if all(value is None or str(value).strip() == "" for value in raw.values()):
            continue
        parsed: dict[str, Any] = {}
        for name, value in raw.items():
            data_value = data_values.get(name)
            parsed[name] = (
                _parse_number(
                    data_value,
                    family="load" if name == "load" else "measurement",
                )
                if _numeric_field(name)
                else data_value
            )
            cells.append(
                RawCell(
                    sheet_name=ws.title,
                    source_row=row_number,
                    source_column=source_letters[name],
                    raw_value=value,
                    parsed_value=parsed[name],
                    canonical_field=name,
                )
            )
            if isinstance(value, str) and value.startswith("=") and data_value is None:
                issues.append(
                    ValidationIssue(
                        "error",
                        "formula_without_cached_value",
                        "Ячейка содержит формулу без сохранённого вычисленного результата.",
                        sheet=ws.title,
                        row=row_number,
                        column=source_letters[name],
                        raw_value=value,
                        suggested_action="Пересчитайте и сохраните книгу в Excel либо замените формулу зафиксированным значением в копии.",
                    )
                )
        if raw.get("stage") is None or str(raw.get("stage")).strip() == "":
            issues.append(
                ValidationIssue(
                    "warning",
                    "missing_stage",
                    "Номер ступени отсутствует; строка сохранена без искусственного номера.",
                    sheet=ws.title,
                    row=row_number,
                    column=source_letters.get("stage"),
                    raw_value=raw.get("stage"),
                    suggested_action="Уточните номер ступени в исходном журнале или mapping-аудите.",
                )
            )
        if parsed.get("load") is None:
            issues.append(
                ValidationIssue(
                    "error",
                    "invalid_load_cell",
                    "Нагрузка не распознана как конечное число.",
                    sheet=ws.title,
                    row=row_number,
                    column=source_letters.get("load"),
                    raw_value=raw.get("load"),
                    suggested_action="Исправьте значение в копии журнала или зарегистрируйте коррекцию.",
                )
            )
        failure_item = next(
            ((name, value) for name, value in raw.items() if _is_failure(value)),
            None,
        )
        failure_source = failure_item[1] if failure_item else None
        missing_markers = {"", "-", "—", "нет", "na", "n/a"}
        for measurement_name in (
            *_MEASUREMENT_COLUMNS,
            "reference_indicator",
            "horizontal_indicator",
        ):
            if measurement_name not in raw:
                continue
            raw_measurement = raw.get(measurement_name)
            normalized_measurement = str(raw_measurement or "").strip().casefold()
            if (
                normalized_measurement not in missing_markers
                and parsed.get(measurement_name) is None
                and not _is_failure(raw_measurement)
            ):
                issues.append(
                    ValidationIssue(
                        "error",
                        "invalid_measurement_cell",
                        f"{measurement_name} не распознан как конечное число или пустое значение.",
                        sheet=ws.title,
                        row=row_number,
                        column=source_letters.get(measurement_name),
                        raw_value=raw_measurement,
                        suggested_action="Проверьте ячейку или зарегистрируйте явную коррекцию.",
                    )
                )
        if failure_source is not None:
            parsed["status"] = "failure"
        indicator_name = next((name for name in _MEASUREMENT_COLUMNS if name in mapping), None)
        record = {name: parsed.get(name) for name in mapping}
        if failure_source is not None and "status" not in record:
            record["status"] = parsed["status"]
        record.update(
            {
                "sheet_name": ws.title,
                "source_row": row_number,
                "source_columns": json.dumps(source_letters, ensure_ascii=False, sort_keys=True),
                "sequence_index": sequence_start + len(records),
                "raw_stage": raw.get("stage"),
                "raw_indicator": raw.get(indicator_name) if indicator_name else None,
                "raw_load": raw.get("load"),
                "parsed_stage": parsed.get("stage"),
                "parsed_indicator": parsed.get(indicator_name) if indicator_name else None,
                "parsed_load": parsed.get("load"),
                "load_unit": load_unit,
                "source_load_unit": load_unit,
                "status_raw": raw.get("status"),
                "failure_marker_raw": failure_source,
                "failure_marker_field": failure_item[0] if failure_item else None,
            }
        )
        records.append(record)
    return records, cells, issues


def _legacy_test_id(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    for pattern in _LEGACY_TITLE_PATTERNS:
        match = pattern.match(value)
        if match:
            return match.group(1)
    return None


def _legacy_header_score(value: Any, role: str) -> int:
    text = _normalise_header(value)
    tokens = {
        "stage": ("шаг", "ступен", "номер"),
        "indicator": ("индикатор", "показан", "ич", "осадк"),
        "load": ("нагруз", "усили", "сила", "f"),
    }
    return sum(token in text for token in tokens[role])


def _parse_legacy_workbook(
    workbook, raw_workbook=None
) -> tuple[pd.DataFrame, list[RawCell], list[ValidationIssue], list[dict[str, Any]]]:
    records: list[dict[str, Any]] = []
    cells: list[RawCell] = []
    issues = [
        ValidationIssue(
            "warning",
            "heuristic_import",
            "Использован legacy heuristic: сопоставление требует инженерной проверки.",
            suggested_action="Сохраните подтвержденный mapping и повторите импорт в interactive режиме.",
        )
    ]
    mappings: list[dict[str, Any]] = []
    seen: set[str] = set()
    for ws in workbook.worksheets:
        raw_ws = raw_workbook[ws.title] if raw_workbook is not None else ws
        titles: list[tuple[int, int, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                test_id = _legacy_test_id(cell.value)
                if test_id:
                    titles.append((cell.row, cell.column, test_id))
                    break
        for position, (title_row, title_col, original_id) in enumerate(titles):
            end_row = titles[position + 1][0] if position + 1 < len(titles) else int(ws.max_row or 0) + 1
            test_id = original_id
            if original_id in seen:
                issues.append(
                    ValidationIssue(
                        "error",
                        "duplicate_legacy_test_id",
                        f"Legacy ID испытания {original_id} повторяется; скрытый суффикс не добавлен.",
                        test_id=original_id,
                        sheet=ws.title,
                        row=title_row,
                        raw_value=original_id,
                        suggested_action="Назначьте уникальный test_id через подтверждённый mapping/паспорт.",
                    )
                )
            seen.add(original_id)
            best: tuple[int, dict[str, int], int] | None = None
            for row_number in range(title_row + 1, min(end_row, title_row + 7)):
                candidate: dict[str, int] = {}
                score = 0
                for role in ("stage", "indicator", "load"):
                    choices = [
                        (cell.column, _legacy_header_score(cell.value, role)) for cell in ws[row_number]
                    ]
                    column, value_score = max(choices, key=lambda item: item[1], default=(0, 0))
                    if value_score:
                        candidate[role] = column
                        score += value_score
                if len(candidate) == 3 and len(set(candidate.values())) == 3 and (best is None or score > best[2]):
                    best = (row_number, candidate, score)
            if best is None:
                issues.append(
                    ValidationIssue(
                        "warning",
                        "legacy_block_unmapped",
                        f"Блок испытания {test_id} не распознан и пропущен.",
                        test_id=test_id,
                        sheet=ws.title,
                        row=title_row,
                        suggested_action="Используйте interactive mapping.",
                    )
                )
                continue
            header_row, legacy_map, _ = best
            mappings.append({"sheet_name": ws.title, "test_id": test_id, "header_row": header_row, "mapping": legacy_map})
            for row_number in range(header_row + 1, end_row):
                raw_stage = _row_value(ws, row_number, legacy_map["stage"])
                raw_indicator = _row_value(ws, row_number, legacy_map["indicator"])
                raw_load = _row_value(ws, row_number, legacy_map["load"])
                original_stage = _row_value(raw_ws, row_number, legacy_map["stage"])
                original_indicator = _row_value(raw_ws, row_number, legacy_map["indicator"])
                original_load = _row_value(raw_ws, row_number, legacy_map["load"])
                if all(
                    value is None or str(value).strip() == ""
                    for value in (original_stage, original_indicator, original_load)
                ):
                    continue
                parsed_load = _parse_number(raw_load, family="load")
                parsed_indicator = _parse_number(raw_indicator, family="measurement")
                failure = _is_failure(raw_indicator) or _is_failure(raw_load)
                if parsed_load is None and not failure:
                    issues.append(
                        ValidationIssue(
                            "warning",
                            "legacy_incomplete_row",
                            "Legacy-строка сохранена, но нагрузка не распознана.",
                            test_id=test_id,
                            sheet=ws.title,
                            row=row_number,
                            raw_value=raw_load,
                            suggested_action="Проверьте исходный журнал.",
                        )
                    )
                letters = {}
                _, get_column_letter = _excel_modules()
                for role, column in legacy_map.items():
                    letters[role] = get_column_letter(column)
                    raw_value = {
                        "stage": original_stage,
                        "indicator": original_indicator,
                        "load": original_load,
                    }[role]
                    cached_value = {
                        "stage": raw_stage,
                        "indicator": raw_indicator,
                        "load": raw_load,
                    }[role]
                    parsed_value = {"stage": _parse_number(raw_stage, family="stage"), "indicator": parsed_indicator, "load": parsed_load}[role]
                    cells.append(RawCell(ws.title, row_number, letters[role], raw_value, parsed_value, role))
                    if isinstance(raw_value, str) and raw_value.startswith("=") and cached_value is None:
                        issues.append(
                            ValidationIssue(
                                "error",
                                "formula_without_cached_value",
                                "Legacy-формула не имеет сохранённого вычисленного значения.",
                                test_id=test_id,
                                sheet=ws.title,
                                row=row_number,
                                column=letters[role],
                                raw_value=raw_value,
                                suggested_action="Пересчитайте и сохраните книгу в Excel/LibreOffice либо замените формулу зафиксированным значением.",
                            )
                        )
                records.append(
                    {
                        "test_id": test_id,
                        "test_id_raw": original_id,
                        "stage": _parse_number(raw_stage, family="stage"),
                        "load": parsed_load,
                        "indicator_1": parsed_indicator,
                        "status": "failure" if failure else None,
                        "sheet_name": ws.title,
                        "source_row": row_number,
                        "source_columns": json.dumps(letters, ensure_ascii=False, sort_keys=True),
                        "sequence_index": len(records),
                        "raw_stage": original_stage,
                        "raw_indicator": original_indicator,
                        "raw_load": original_load,
                        "parsed_stage": _parse_number(raw_stage, family="stage"),
                        "parsed_indicator": parsed_indicator,
                        "parsed_load": parsed_load,
                        "load_unit": _infer_load_unit(_row_value(raw_ws, header_row, legacy_map["load"])),
                        "source_load_unit": _infer_load_unit(_row_value(raw_ws, header_row, legacy_map["load"])),
                        "indicator_requires_calibration": True,
                        "status_raw": None,
                        "failure_marker_raw": (
                            raw_indicator if _is_failure(raw_indicator) else raw_load if _is_failure(raw_load) else None
                        ),
                        "failure_marker_field": (
                            "indicator_1" if _is_failure(raw_indicator) else "load" if _is_failure(raw_load) else None
                        ),
                    }
                )
    if records:
        issues.append(
            ValidationIssue(
                "warning",
                "legacy_indicator_calibration_required",
                "Показания legacy-индикатора импортированы как raw; направление и развёртка шкалы должны быть подтверждены в следующем этапе.",
                suggested_action="Укажите паспорт индикатора и проверьте переходы шкалы.",
            )
        )
    return pd.DataFrame(records), cells, issues, mappings


def _xlsx_failure_result(
    data: bytes, import_mode: str, *, code: str, message: str
) -> ProtocolImportResult:
    issue = ValidationIssue(
        "error",
        code,
        message,
        suggested_action="Проверьте источник файла и экспортируйте книгу заново без внешних связей/макросов.",
    )
    return ProtocolImportResult(
        pd.DataFrame(),
        {
            "format": "xlsx",
            "import_mode": import_mode,
            "input_file_sha256": sha256_bytes(data),
            "rows": 0,
            "columns": [],
            "sheets": [],
            "raw_cell_count": 0,
            "blocking_issue_count": 1,
        },
        [issue],
        pd.DataFrame(),
    )


def read_protocol_excel(
    source: str | Path | bytes | bytearray | BinaryIO,
    *,
    import_mode: str = "strict",
    column_mapping: dict[str, str | int] | None = None,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> ProtocolImportResult:
    if import_mode not in IMPORT_MODES:
        raise ValueError(f"import_mode должен быть одним из {IMPORT_MODES}.")
    if import_mode == "interactive" and not column_mapping:
        issue = ValidationIssue(
            "error",
            "interactive_mapping_required",
            "Для interactive импорта требуется сохраненное сопоставление столбцов.",
            suggested_action="Выберите столбцы в интерфейсе или передайте mapping JSON.",
        )
        return ProtocolImportResult(pd.DataFrame(), {"format": "xlsx", "import_mode": import_mode}, [issue], pd.DataFrame())
    data = _read_bytes(source)
    workbook_values = None
    workbook_raw = None
    try:
        _validate_xlsx_archive(data)
        load_workbook, _ = _excel_modules()
        workbook_values = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False
        )
        try:
            workbook_raw = load_workbook(
                io.BytesIO(data), read_only=True, data_only=False, keep_links=False
            )
        except Exception:
            workbook_values.close()
            raise
        _validate_workbook_limits(workbook_raw)
    except _XlsxGuardError as exc:
        if workbook_values is not None:
            workbook_values.close()
        if workbook_raw is not None:
            workbook_raw.close()
        return _xlsx_failure_result(data, import_mode, code=exc.code, message=str(exc))
    except Exception as exc:
        if workbook_values is not None:
            workbook_values.close()
        if workbook_raw is not None:
            workbook_raw.close()
        return _xlsx_failure_result(
            data,
            import_mode,
            code=_xlsx_exception_code(exc),
            message=f"Не удалось безопасно разобрать XLSX: {exc}",
        )
    all_records: list[dict[str, Any]] = []
    all_cells: list[RawCell] = []
    all_issues: list[ValidationIssue] = []
    sheet_info: list[dict[str, Any]] = []
    try:
        selected = [sheet_name] if sheet_name else workbook_raw.sheetnames
        for name in selected:
            if name not in workbook_raw.sheetnames:
                all_issues.append(
                    ValidationIssue("error", "missing_sheet", f"Лист {name!r} отсутствует.", sheet=name)
                )
                continue
            ws = workbook_values[name]
            raw_ws = workbook_raw[name]
            discovery = _discover_header(
                raw_ws,
                header_row,
                allow_unrecognized=import_mode == "interactive",
            )
            if discovery is None:
                explicitly_selected = sheet_name is not None
                suspicious_table = _looks_like_tabular_sheet(raw_ws)
                level = (
                    "warning"
                    if import_mode == "heuristic"
                    else "error" if explicitly_selected or suspicious_table else "info"
                )
                all_issues.append(
                    ValidationIssue(
                        level,
                        (
                            "schema_not_recognized"
                            if explicitly_selected or suspicious_table
                            else "sheet_skipped_no_schema"
                        ),
                        (
                            "Не найдена однозначная строка заголовков long-table."
                            if explicitly_selected or suspicious_table
                            else "Служебный лист без протокольной схемы пропущен."
                        ),
                        sheet=name,
                        suggested_action="Выберите лист/строку заголовков или heuristic legacy режим.",
                    )
                )
                continue
            mapping, letters, mapping_issues = _resolve_mapping(
                discovery,
                mode=import_mode,
                supplied=column_mapping,
                sheet_name=name,
            )
            all_issues.extend(mapping_issues)
            sheet_info.append(
                {
                    "sheet_name": name,
                    "header_row": discovery["header_row"],
                    "mapping": letters,
                }
            )
            if any(item.blocks_processing for item in mapping_issues):
                continue
            records, cells, row_issues = _parse_long_table_sheet(
                ws,
                discovery,
                mapping,
                letters,
                sequence_start=len(all_records),
                raw_ws=raw_ws,
            )
            all_records.extend(records)
            all_cells.extend(cells)
            all_issues.extend(row_issues)

        if not all_records and import_mode == "heuristic":
            legacy_frame, legacy_cells, legacy_issues, legacy_mappings = _parse_legacy_workbook(
                workbook_values, workbook_raw
            )
            if not legacy_frame.empty:
                all_records = legacy_frame.to_dict(orient="records")
                all_cells.extend(legacy_cells)
                all_issues = [item for item in all_issues if item.code == "missing_sheet"]
                all_issues.extend(legacy_issues)
                sheet_info = legacy_mappings
        frame = pd.DataFrame(all_records)
        raw_cells = pd.DataFrame([item.to_dict() for item in all_cells])
        if not frame.empty and {"test_id", "sheet_name"}.issubset(frame.columns):
            sheet_counts = frame.groupby(frame["test_id"].astype(str))["sheet_name"].nunique()
            for duplicate_id in sheet_counts[sheet_counts > 1].index:
                duplicate_sheets = sorted(
                    frame.loc[frame["test_id"].astype(str) == duplicate_id, "sheet_name"]
                    .astype(str)
                    .unique()
                )
                all_issues.append(
                    ValidationIssue(
                        "error",
                        "duplicate_test_id_across_sheets",
                        f"ID испытания {duplicate_id} повторяется на листах: "
                        + ", ".join(duplicate_sheets),
                        test_id=str(duplicate_id),
                        raw_value=duplicate_sheets,
                        suggested_action="Назначьте уникальные test_id; автоматический суффикс не добавляется.",
                    )
                )
        if not all_records:
            all_issues.append(
                ValidationIssue(
                    "error",
                    "no_protocol_rows",
                    "В книге не найдено ни одной строки протокола.",
                    suggested_action="Проверьте лист, заголовок и режим импорта.",
                )
            )
        if import_mode == "heuristic" and all_records and not any(item.code == "heuristic_import" for item in all_issues):
            all_issues.append(
                ValidationIssue(
                    "warning",
                    "heuristic_import",
                    "Импорт выполнен в heuristic режиме; mapping требует ручного подтверждения.",
                    suggested_action="Сохраните mapping и повторите импорт в interactive режиме.",
                )
            )
        info = {
            "format": "xlsx",
            "import_mode": import_mode,
            "input_file_sha256": sha256_bytes(data),
            "rows": len(frame),
            "columns": list(frame.columns),
            "sheets": sheet_info,
            "raw_cell_count": len(raw_cells),
            "blocking_issue_count": sum(bool(item.blocks_processing) for item in all_issues),
        }
        return ProtocolImportResult(frame, info, all_issues, raw_cells)
    finally:
        workbook_values.close()
        workbook_raw.close()


def read_protocol(
    source: str | Path | bytes | bytearray | BinaryIO,
    *,
    filename: str | None = None,
    import_mode: str = "strict",
    column_mapping: dict[str, str | int] | None = None,
    sheet_name: str | None = None,
    header_row: int | None = None,
) -> ProtocolImportResult:
    name = filename or (str(source) if isinstance(source, (str, Path)) else "")
    suffix = Path(name).suffix.casefold()
    data = _read_bytes(source)
    if suffix in {".xlsx", ".xlsm"} or data.startswith(b"PK\x03\x04"):
        result = read_protocol_excel(
            data,
            import_mode=import_mode,
            column_mapping=column_mapping,
            sheet_name=sheet_name,
            header_row=header_row,
        )
        if suffix == ".xlsm":
            result.issues.append(
                ValidationIssue(
                    "warning",
                    "xlsm_macros_ignored",
                    "Книга XLSM прочитана только как данные; VBA-макросы не загружались и не выполнялись.",
                    suggested_action="Для архивного протокола предпочтителен экспорт в XLSX без макросов.",
                )
            )
            result.info["macros_ignored"] = True
        result.info["blocking_issue_count"] = sum(
            bool(item.blocks_processing) for item in result.issues
        )
        return result
    frame, info = read_protocol_csv(data)
    info["import_mode"] = import_mode
    allowed = set(REQUIRED_COLUMNS) | set(OPTIONAL_PROTOCOL_COLUMNS) | {"pair_id"}
    issues: list[ValidationIssue] = []
    for column in info.get("source_columns", []):
        if column not in allowed:
            issues.append(
                ValidationIssue(
                    "error" if import_mode in {"strict", "interactive"} else "warning",
                    "unknown_header",
                    f"Неизвестный заголовок CSV {column!r}.",
                    sheet="CSV",
                    row=1,
                    column=column,
                    raw_value=column,
                    suggested_action="Используйте каноническое имя столбца или XLSX interactive mapping.",
                )
            )
    for required in REQUIRED_COLUMNS:
        if required not in info.get("source_columns", []):
            issues.append(
                ValidationIssue(
                    "error",
                    "missing_required_mapping",
                    f"Нет обязательного CSV-столбца {required}.",
                    sheet="CSV",
                    row=1,
                    column=required,
                    suggested_action="Добавьте канонический заголовок.",
                )
            )
    if not any(name in info.get("source_columns", []) for name in _MEASUREMENT_COLUMNS):
        issues.append(
            ValidationIssue(
                "error",
                "missing_measurement_mapping",
                "Нет settlement или indicator_1..4.",
                sheet="CSV",
                row=1,
                suggested_action="Добавьте источник осадки.",
            )
        )
    return ProtocolImportResult(frame, info, issues, pd.DataFrame())
