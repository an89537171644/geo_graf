"""Deterministic, review-oriented XLSX report rendering.

The workbook is an exported view of the scientific results: it deliberately
contains no executable Excel formula cells.  Scientific formulae are recorded
as literal text in ``Methodology`` and the exact exported table ranges are
listed there as well.

Excel does not permit ``/`` in a sheet title.  The public sheet-name mapping is
therefore explicit and is also written to ``Methodology``::

    Failure/censoring -> Failure censoring

All other logical names currently map to themselves.  Every name is at most
31 characters and is safe for Excel.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import re
import zipfile
from collections.abc import Mapping, Sequence
from datetime import date, datetime, time
from decimal import Decimal, ROUND_HALF_UP
from numbers import Real
from pathlib import Path
from typing import Any
from urllib.parse import unquote

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEET_NAME_MAP: dict[str, str] = {
    "Project passport": "Project passport",
    "Raw data": "Raw data",
    "Prepared data": "Prepared data",
    "Indicator passports": "Indicator passports",
    "Indicator audit": "Indicator audit",
    "QC issues": "QC issues",
    "Failure/censoring": "Failure censoring",
    "pcr": "pcr",
    "Moduli": "Moduli",
    "Group comparison": "Group comparison",
    "Plots index": "Plots index",
    "Audit trail": "Audit trail",
    "Provenance": "Provenance",
    "Methodology": "Methodology",
}

_FIXED_TIMESTAMP = datetime(2000, 1, 1, 0, 0, 0)
_ZIP_TIMESTAMP = (2000, 1, 1, 0, 0, 0)
_HEADER_ROW = 3
_DATA_ROW = 4
_DANGEROUS_PREFIXES = ("=", "+", "-", "@")
_MACHINE_NUMBER_FORMAT = "0.###############"
_MEDIA_TYPES = {
    ".csv": "text/csv",
    ".html": "text/html",
    ".jpeg": "image/jpeg",
    ".jpg": "image/jpeg",
    ".json": "application/json",
    ".md": "text/markdown",
    ".pdf": "application/pdf",
    ".png": "image/png",
    ".svg": "image/svg+xml",
    ".txt": "text/plain",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".zip": "application/zip",
}
_METHODOLOGY_HEADERS = [
    "record_type",
    "record_id",
    "name",
    "value",
    "expression",
    "scope",
    "units",
    "range",
    "artifact_path",
    "artifact_href",
    "artifact_sha256",
    "artifact_bytes",
    "media_type",
    "notes",
]


def _nonfinite_token(value: Any) -> str | None:
    """Return a sign-preserving literal for a non-finite numeric value."""

    if isinstance(value, Decimal):
        if value.is_nan():
            return "NaN"
        if value.is_infinite():
            return "-Inf" if value.is_signed() else "+Inf"
        return None
    if isinstance(value, Real) and not isinstance(value, bool):
        try:
            numeric = float(value)
        except (TypeError, ValueError, OverflowError):
            return None
        if math.isnan(numeric):
            return "NaN"
        if math.isinf(numeric):
            return "-Inf" if numeric < 0 else "+Inf"
    return None


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if _nonfinite_token(value) is not None:
        # NaN and infinities are scientific states, not empty spreadsheet cells.
        return False
    if isinstance(value, (list, tuple, dict, set, bytes, bytearray, memoryview)):
        return False
    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        return False
    return bool(missing) if isinstance(missing, (bool, type(pd.NA))) else False


def _json_text(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )


def _literal_payload(value: Any) -> Any:
    """Convert nested/table-like values to deterministic JSON-compatible data."""

    nonfinite = _nonfinite_token(value)
    if nonfinite is not None:
        return nonfinite
    if isinstance(value, pd.DataFrame):
        return {
            "columns": [str(column) for column in value.columns],
            "data": [
                [_literal_payload(item) for item in row]
                for row in value.itertuples(index=False, name=None)
            ],
        }
    if isinstance(value, pd.Series):
        return [_literal_payload(item) for item in value.tolist()]
    if isinstance(value, Mapping):
        return {
            str(key): _literal_payload(value[key])
            for key in sorted(value, key=str)
        }
    if isinstance(value, set):
        return [_literal_payload(item) for item in sorted(value, key=str)]
    if isinstance(value, (list, tuple)):
        return [_literal_payload(item) for item in value]
    if hasattr(value, "to_dict"):
        try:
            return _literal_payload(value.to_dict())
        except (TypeError, ValueError):
            pass
    if _is_missing(value):
        return None
    if hasattr(value, "item") and not isinstance(value, (str, bytes, bytearray)):
        try:
            return _literal_payload(value.item())
        except (AttributeError, TypeError, ValueError):
            pass
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (Decimal, Path)):
        return str(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _literal_text(value: Any) -> str:
    return _json_text(_literal_payload(value))


def _excel_value(value: Any) -> Any:
    """Return a supported value without coercing source strings to numbers."""

    nonfinite = _nonfinite_token(value)
    if nonfinite is not None:
        return nonfinite
    if _is_missing(value):
        return None
    if hasattr(value, "item") and not isinstance(value, (str, bytes, bytearray)):
        try:
            value = value.item()
        except (AttributeError, TypeError, ValueError):
            pass
    if isinstance(value, (Mapping, list, tuple, set, pd.DataFrame, pd.Series)):
        return _literal_text(value)
    if isinstance(value, datetime):
        # openpyxl rejects timezone-aware datetimes.  ISO text is lossless and
        # avoids locale/timezone dependent workbook serialisation.
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _set_literal(cell: Cell, value: Any) -> None:
    """Write strings as literal text, including formula-like user input."""

    converted = _excel_value(value)
    cell.value = converted
    if isinstance(converted, str):
        # Setting data_type after value assignment prevents ``=...`` from
        # becoming an OOXML formula while preserving the exact visible text.
        cell.data_type = "s"
        if converted.startswith(_DANGEROUS_PREFIXES):
            cell.quotePrefix = True


def _object_mapping(value: Any) -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, Mapping):
        return dict(value)
    if hasattr(value, "to_dict"):
        payload = value.to_dict()
        if isinstance(payload, Mapping):
            return dict(payload)
    if hasattr(value, "__dict__"):
        return dict(vars(value))
    return {"value": value}


def _unique_headers(values: Sequence[Any]) -> list[str]:
    result: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value) if value is not None and str(value) else f"column_{index}"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}__{counts[base]}")
    return result


def _tabular(value: Any) -> tuple[list[str], list[list[Any]]]:
    if value is None:
        return [], []
    if isinstance(value, pd.DataFrame):
        headers = _unique_headers(list(value.columns))
        rows = [list(row) for row in value.itertuples(index=False, name=None)]
        return headers, rows
    if isinstance(value, Mapping):
        mapping = dict(value)
        if not mapping:
            return [], []
        return _unique_headers(list(mapping)), [list(mapping.values())]
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        items = list(value)
        if not items:
            return [], []
        mappings = [_object_mapping(item) for item in items]
        headers: list[str] = []
        for item in mappings:
            for key in item:
                text = str(key)
                if text not in headers:
                    headers.append(text)
        return _unique_headers(headers), [[item.get(name) for name in headers] for item in mappings]
    mapping = _object_mapping(value)
    return _unique_headers(list(mapping)), [list(mapping.values())]


def _keyed_results(value: Any, id_name: str = "test_id") -> tuple[list[str], list[list[Any]]]:
    if not isinstance(value, Mapping):
        return _tabular(value)
    rows: list[dict[str, Any]] = []
    for key in sorted(value, key=str):
        payload = _object_mapping(value[key])
        payload = {id_name: key, **payload}
        rows.append(payload)
    return _tabular(rows)


def _combined_tables(value: Any) -> tuple[list[str], list[list[Any]]]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes, bytearray)):
        return _tabular(value)
    frames = list(value)
    if not frames:
        return [], []
    records: list[dict[str, Any]] = []
    for index, item in enumerate(frames, start=1):
        headers, rows = _tabular(item)
        for row in rows:
            records.append({"comparison_index": index, **dict(zip(headers, row, strict=False))})
    return _tabular(records)


def _flatten_mapping(value: Any, prefix: str = "") -> list[dict[str, Any]]:
    mapping = _object_mapping(value)
    rows: list[dict[str, Any]] = []
    for key in sorted(mapping, key=str):
        path = f"{prefix}.{key}" if prefix else str(key)
        item = mapping[key]
        if isinstance(item, Mapping):
            rows.extend(_flatten_mapping(item, path))
        else:
            rows.append(
                {
                    "field": path,
                    "value": _literal_text(item) if isinstance(item, (list, tuple, set)) else item,
                    "value_type": type(item).__name__,
                }
            )
    return rows


def _result_table_value(
    result_tables: Mapping[str, Any] | None,
    names: Sequence[str],
) -> Any:
    for name in names:
        if isinstance(result_tables, Mapping) and name in result_tables:
            return result_tables[name]
    return None


def _first_not_none(*values: Any) -> Any:
    for value in values:
        if value is not None:
            return value
    return None


def _display_value(value: Any, digits: int) -> str | None:
    nonfinite = _nonfinite_token(value)
    if nonfinite is not None:
        return nonfinite
    if _is_missing(value):
        return None
    if isinstance(value, bool):
        return str(value)
    try:
        quantizer = Decimal(1).scaleb(-int(digits))
        rounded = Decimal(str(value)).quantize(quantizer, rounding=ROUND_HALF_UP)
    except (ArithmeticError, TypeError, ValueError):
        return str(value)
    return f"{rounded:.{int(digits)}f}".replace(".", ",")


def _machine_repr(value: Any) -> tuple[bool, str | None]:
    """Return an exact textual companion for a machine numeric value."""

    nonfinite = _nonfinite_token(value)
    if nonfinite is not None:
        return True, nonfinite
    if _is_missing(value):
        return False, None
    if hasattr(value, "item") and not isinstance(value, (str, bytes, bytearray)):
        try:
            value = value.item()
        except (AttributeError, TypeError, ValueError):
            pass
    if isinstance(value, bool):
        return False, None
    if isinstance(value, Decimal):
        return True, str(value)
    if isinstance(value, int):
        return True, str(value)
    if isinstance(value, float):
        return True, repr(value)
    return False, None


def _with_exact_machine_columns(
    logical_name: str,
    headers: list[str],
    rows: list[list[Any]],
) -> tuple[list[str], list[list[Any]], list[dict[str, Any]]]:
    """Add lossless text companions without replacing numeric convenience cells."""

    if logical_name == "Raw data" or not headers or not rows:
        return headers, rows, []
    result_headers = list(headers)
    result_rows = [list(row) for row in rows]
    records: list[dict[str, Any]] = []
    for column_index, column in enumerate(headers):
        representations = [
            _machine_repr(row[column_index]) if column_index < len(row) else (False, None)
            for row in rows
        ]
        if not any(is_numeric for is_numeric, _ in representations):
            continue
        exact_name = f"{column}__exact_machine"
        if exact_name in result_headers:
            # An explicit caller-supplied companion remains authoritative.
            continue
        result_headers.append(exact_name)
        for row, (is_numeric, representation) in zip(
            result_rows,
            representations,
            strict=True,
        ):
            row.append(representation if is_numeric else None)
        records.append(
            {
                "record_type": "machine_exact_repr",
                "record_id": f"{logical_name}.{column}",
                "name": exact_name,
                "scope": logical_name,
                "notes": (
                    "Exact source numeric representation as literal text; the source column "
                    "remains numeric where OOXML supports it."
                ),
            }
        )
    return result_headers, result_rows, records


def _with_display_columns(
    logical_name: str,
    headers: list[str],
    rows: list[list[Any]],
    display_rounding: Mapping[str, Any] | None,
) -> tuple[list[str], list[list[Any]], list[dict[str, Any]]]:
    if not display_rounding:
        return headers, rows, []
    requested = display_rounding.get(logical_name)
    if requested is None:
        requested = display_rounding.get(SHEET_NAME_MAP[logical_name])
    if requested is None and logical_name != "Raw data":
        default_digits = display_rounding.get("default")
        if isinstance(default_digits, int) and not isinstance(default_digits, bool):
            requested = {
                header: default_digits
                for column_index, header in enumerate(headers)
                if any(
                    column_index < len(row)
                    and isinstance(row[column_index], (int, float, Decimal))
                    and not isinstance(row[column_index], bool)
                    for row in rows
                )
            }
    if not requested:
        return headers, rows, []
    if not isinstance(requested, Mapping):
        raise ValueError(f"Display-rounding specification for {logical_name} must be a mapping.")
    result_headers = list(headers)
    result_rows = [list(row) for row in rows]
    records: list[dict[str, Any]] = []
    for column, raw_digits in requested.items():
        if column not in headers:
            raise ValueError(f"Display column source not found in {logical_name}: {column}")
        digits = int(raw_digits)
        if digits < 0 or digits > 15:
            raise ValueError(f"Unsupported display precision for {logical_name}.{column}: {digits}")
        source_index = headers.index(column)
        display_name = f"{column}__display"
        if display_name in result_headers:
            raise ValueError(f"Display column already exists: {logical_name}.{display_name}")
        result_headers.append(display_name)
        for row in result_rows:
            row.append(_display_value(row[source_index], digits))
        records.append(
            {
                "record_type": "rounding",
                "record_id": f"{logical_name}.{column}",
                "name": display_name,
                "value": digits,
                "scope": logical_name,
                "notes": "Display-only decimal places; source machine column is unchanged.",
            }
        )
    return result_headers, result_rows, records


def _artifact_items(value: Any, *, source_is_plot: bool) -> list[dict[str, Any]]:
    if value is None:
        return []
    if isinstance(value, Mapping):
        items = []
        for name, record in value.items():
            if isinstance(record, Mapping):
                items.append({"name": name, **dict(record)})
            else:
                items.append({"name": name, "payload": record})
    elif isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        items = [_object_mapping(record) for record in value]
    else:
        items = [{"name": "artifact", "payload": value}]

    result: list[dict[str, Any]] = []
    for index, source in enumerate(items, start=1):
        name = str(
            source.get("name")
            or source.get("artifact_name")
            or source.get("path")
            or f"artifact_{index}"
        )
        path_value = source.get("path")
        href = source.get("href") or source.get("url") or path_value or name
        payload = source.get("payload", source.get("content"))
        byte_count = source.get("bytes") if isinstance(source.get("bytes"), int) else None
        if payload is None and isinstance(source.get("bytes"), (bytes, bytearray, memoryview)):
            payload = source.get("bytes")
        if payload is None and isinstance(path_value, (str, Path)):
            candidate = Path(path_value)
            if candidate.is_file():
                payload = candidate.read_bytes()
        if isinstance(payload, str):
            payload = payload.encode("utf-8")
        elif isinstance(payload, memoryview):
            payload = payload.tobytes()
        elif isinstance(payload, bytearray):
            payload = bytes(payload)
        digest = source.get("sha256")
        if isinstance(payload, bytes):
            actual_digest = hashlib.sha256(payload).hexdigest()
            if digest and str(digest).casefold() != actual_digest:
                raise ValueError(f"SHA-256 mismatch for artifact {name}")
            digest = actual_digest
            byte_count = len(payload)
        elif digest and not str(digest).startswith("excluded:"):
            digest_text = str(digest).casefold()
            if len(digest_text) != 64 or any(
                character not in "0123456789abcdef" for character in digest_text
            ):
                raise ValueError(f"Invalid SHA-256 for artifact {name}")
        suffix = Path(str(path_value or name)).suffix.casefold()
        media_type = source.get("media_type") or _MEDIA_TYPES.get(suffix)
        is_plot = bool(
            source_is_plot
            or str(media_type or "").casefold().startswith("image/")
            or str(media_type or "").casefold() == "application/pdf"
            or suffix in {".svg", ".png", ".jpg", ".jpeg", ".pdf"}
        )
        result.append(
            {
                "artifact_name": name,
                "path": str(path_value or name),
                "href": str(href),
                "bytes": byte_count,
                "sha256": str(digest) if digest else None,
                "media_type": str(media_type or "application/octet-stream"),
                "is_plot": is_plot,
                "manifest_status": (
                    "excluded_by_policy"
                    if str(digest or "").startswith("excluded:")
                    else "verified"
                    if digest
                    else "review_required"
                ),
            }
        )
    return result


def _artifact_manifest(artifacts: Any, plots: Any) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for record in [*_artifact_items(artifacts, source_is_plot=False), *_artifact_items(plots, source_is_plot=True)]:
        name = record["artifact_name"]
        existing = merged.get(name)
        if existing is not None:
            old_hash = existing.get("sha256")
            new_hash = record.get("sha256")
            if old_hash and new_hash and old_hash != new_hash:
                raise ValueError(f"Conflicting artifact records for {name}")
            existing.update({key: value for key, value in record.items() if value is not None})
            existing["is_plot"] = bool(existing.get("is_plot") or record.get("is_plot"))
        else:
            merged[name] = record
    return [merged[name] for name in sorted(merged, key=str.casefold)]


def _formula_records(formulas: Any) -> list[dict[str, Any]]:
    if formulas is None:
        return [
            {
                "record_type": "formula_policy",
                "record_id": "xlsx_formula_cells",
                "name": "Executable spreadsheet formula cells",
                "value": 0,
                "notes": "Scientific results are exported values; no hidden Excel recalculation.",
            }
        ]
    if isinstance(formulas, Mapping):
        items = []
        for name in sorted(formulas, key=str):
            value = formulas[name]
            formula_fields = {
                "record_id",
                "id",
                "name",
                "value",
                "expression",
                "formula",
                "scope",
                "units",
                "range",
                "notes",
                "description",
            }
            if isinstance(value, Mapping) and formula_fields.intersection(value):
                items.append({"record_id": name, **dict(value)})
            elif isinstance(value, str):
                items.append({"record_id": name, "expression": value})
            else:
                items.append({"record_id": name, "value": _literal_text(value)})
    else:
        items = [_object_mapping(item) for item in formulas]
    records: list[dict[str, Any]] = []
    for index, item in enumerate(items, start=1):
        record_id = item.get("record_id")
        if record_id is None:
            record_id = item.get("id")
        expression = item.get("expression")
        if expression is None:
            expression = item.get("formula")
        notes = item.get("notes")
        if notes is None:
            notes = item.get("description")
        records.append(
            {
                "record_type": "formula",
                "record_id": record_id if record_id is not None else f"formula_{index}",
                "name": item.get("name"),
                "value": item.get("value"),
                "expression": expression,
                "scope": item.get("scope"),
                "units": item.get("units"),
                "range": item.get("range"),
                "notes": notes,
            }
        )
    return records


def _methodology_records(methodology: Any) -> list[dict[str, Any]]:
    if methodology is None:
        return []
    if isinstance(methodology, Mapping):
        items = [
            {"record_id": key, "value": methodology[key]}
            for key in sorted(methodology, key=str)
        ]
    elif isinstance(methodology, Sequence) and not isinstance(
        methodology, (str, bytes, bytearray)
    ):
        items = [_object_mapping(item) for item in methodology]
    else:
        items = [{"record_id": "methodology", "value": methodology}]
    return [
        {
            "record_type": "methodology",
            "record_id": item.get("record_id") or item.get("id") or f"item_{index}",
            "name": item.get("name"),
            "value": item.get("value"),
            "expression": item.get("expression"),
            "scope": item.get("scope"),
            "units": item.get("units"),
            "range": item.get("range"),
            "notes": item.get("notes") or item.get("description"),
        }
        for index, item in enumerate(items, start=1)
    ]


def _contains_review_required(value: Any) -> bool:
    if isinstance(value, Mapping):
        return any(_contains_review_required(item) for item in value.values())
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        return any(_contains_review_required(item) for item in value)
    return isinstance(value, str) and "review_required" in value.casefold()


def _review_registry_records(value: Any) -> list[dict[str, Any]]:
    if value is None or isinstance(value, bool):
        return []
    if isinstance(value, Mapping):
        items = [dict(value)]
    elif isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
        items = [
            _object_mapping(item) if not isinstance(item, str) else {"message": item}
            for item in value
        ]
    else:
        items = [{"message": str(value)}]
    records = []
    for index, item in enumerate(items, start=1):
        message = item.get("message")
        records.append(
            {
                "record_type": "review_required",
                "record_id": item.get("record_id") or item.get("id") or f"review_{index:04d}",
                "name": item.get("category") or "engineering_review",
                "value": item.get("status") or "review_required",
                "scope": item.get("scope") or item.get("test_id") or "project",
                "notes": message if message is not None else _literal_text(item),
            }
        )
    return records


def _range_for(headers: list[str], rows: list[list[Any]]) -> str:
    last_column = get_column_letter(max(1, len(headers)))
    last_row = max(_HEADER_ROW, _HEADER_ROW + len(rows))
    return f"$A${_HEADER_ROW}:${last_column}${last_row}"


def _safe_package_hyperlink(value: Any) -> bool:
    """Allow only unambiguous relative paths inside the report package."""

    if not isinstance(value, str) or not value:
        return False
    candidate = value
    for _ in range(5):
        if any(ord(character) < 32 for character in candidate):
            return False
        if re.search(r"%(?![0-9A-Fa-f]{2})", candidate):
            return False
        try:
            decoded = unquote(candidate, errors="strict")
        except (UnicodeDecodeError, ValueError):
            return False
        if decoded == candidate:
            break
        candidate = decoded
    else:
        # Deeply nested encodings are ambiguous and therefore not allowlisted.
        return False
    if (
        not candidate
        or candidate.startswith(("/", "\\"))
        or "\\" in candidate
        or ":" in candidate
        or "?" in candidate
        or "#" in candidate
        or any(ord(character) < 32 for character in candidate)
    ):
        return False
    parts = candidate.split("/")
    return bool(parts) and all(part not in {"", ".", ".."} for part in parts)


def _normalise_zip(payload: bytes) -> bytes:
    """Normalise OOXML member order and timestamps for byte reproducibility."""

    source_buffer = io.BytesIO(payload)
    output = io.BytesIO()
    with zipfile.ZipFile(source_buffer, "r") as source, zipfile.ZipFile(
        output,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as target:
        for name in sorted(source.namelist()):
            member_payload = source.read(name)
            if name == "docProps/core.xml":
                member_payload = re.sub(
                    rb"(<dcterms:created\b[^>]*>)[^<]*(</dcterms:created>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    member_payload,
                )
                member_payload = re.sub(
                    rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    member_payload,
                )
            info = zipfile.ZipInfo(filename=name, date_time=_ZIP_TIMESTAMP)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = 0
            target.writestr(
                info,
                member_payload,
                compress_type=zipfile.ZIP_DEFLATED,
                compresslevel=9,
            )
    return output.getvalue()


def build_xlsx_report_package(
    *,
    metadata: Mapping[str, Any] | None = None,
    raw: Any = None,
    prepared: Any = None,
    indicator_passports: Any = None,
    indicator_audit: Any = None,
    qc_issues: Any = None,
    failures: Any = None,
    pcr_results: Any = None,
    moduli: Any = None,
    group_comparisons: Any = None,
    plots: Any = None,
    artifacts: Any = None,
    audit: Any = None,
    provenance: Any = None,
    methodology: Any = None,
    formulas: Any = None,
    display_rounding: Mapping[str, Any] | None = None,
    review_required: Any = None,
    result_tables: Mapping[str, Any] | None = None,
    title: str | None = None,
) -> bytes:
    """Build a deterministic XLSX report and return its complete byte payload.

    ``raw`` strings and ``None`` values are written literally.  Numeric values
    in machine tables remain numeric and are never replaced by rounded text.
    Optional ``display_rounding`` adds explicit ``__display`` columns.

    ``artifacts`` and ``plots`` accept either mappings or record lists.  A
    record may contain ``path``, ``href``, ``payload``/``content``, ``bytes``
    (a byte count or payload), ``sha256`` and ``media_type``.  The complete
    manifest is written to ``Methodology``; plot-like records are additionally
    indexed on ``Plots index``.
    """

    metadata_payload = dict(metadata or {})
    prepared_attrs = getattr(prepared, "attrs", {}) if prepared is not None else {}
    indicator_passports = _first_not_none(
        indicator_passports,
        prepared_attrs.get("indicator_calibration_parameters"),
        _result_table_value(
            result_tables,
            ("indicator_passports", "indicator_calibration_parameters"),
        ),
    )
    indicator_audit = _first_not_none(
        indicator_audit,
        prepared_attrs.get("indicator_processing_audit"),
        _result_table_value(
            result_tables,
            ("indicator_audit", "indicator_processing_audit"),
        ),
    )
    qc_issues = qc_issues if qc_issues is not None else _result_table_value(
        result_tables, ("qc_issues", "validation_issues")
    )
    failures = failures if failures is not None else _result_table_value(
        result_tables, ("failures", "failure_summary")
    )
    pcr_results = pcr_results if pcr_results is not None else _result_table_value(
        result_tables, ("pcr", "pcr_results")
    )
    moduli = moduli if moduli is not None else _result_table_value(
        result_tables, ("moduli", "modulus_results")
    )
    group_comparisons = (
        group_comparisons
        if group_comparisons is not None
        else _result_table_value(result_tables, ("group_comparison", "group_comparisons"))
    )

    manifest = _artifact_manifest(artifacts, plots)
    plot_rows = [record for record in manifest if record["is_plot"]]
    project_rows = _flatten_mapping(metadata_payload)
    project_rows.insert(
        0,
        {
            "field": "report_title",
            "value": title or metadata_payload.get("project") or "Soil Stamp Antonov",
            "value_type": "str",
        },
    )

    tables: dict[str, tuple[list[str], list[list[Any]]]] = {
        "Project passport": _tabular(project_rows),
        "Raw data": _tabular(raw),
        "Prepared data": _tabular(prepared),
        "Indicator passports": _tabular(indicator_passports),
        "Indicator audit": _tabular(indicator_audit),
        "QC issues": _tabular(qc_issues),
        "Failure/censoring": _tabular(failures),
        "pcr": _keyed_results(pcr_results),
        "Moduli": _tabular(moduli),
        "Group comparison": _combined_tables(group_comparisons),
        "Plots index": _tabular(plot_rows),
        "Audit trail": _tabular(
            audit.events if hasattr(audit, "events") else audit
        ),
        "Provenance": _tabular(_flatten_mapping(_object_mapping(provenance))),
    }

    exact_machine_records: list[dict[str, Any]] = []
    rounding_records: list[dict[str, Any]] = []
    for logical_name in list(tables):
        headers, rows = tables[logical_name]
        headers, rows, exact_records = _with_exact_machine_columns(
            logical_name,
            headers,
            rows,
        )
        headers, rows, records = _with_display_columns(
            logical_name,
            headers,
            rows,
            display_rounding,
        )
        tables[logical_name] = headers, rows
        exact_machine_records.extend(exact_records)
        rounding_records.extend(records)

    review_registry_records = _review_registry_records(review_required)
    detected_review = _contains_review_required(
        [metadata_payload, *[row for _, rows in tables.values() for row in rows], manifest]
    )
    explicit_review = review_required is True or bool(review_registry_records)
    effective_review = bool(detected_review or explicit_review)

    methodology_records: list[dict[str, Any]] = []
    for index, (logical_name, excel_name) in enumerate(SHEET_NAME_MAP.items(), start=1):
        methodology_records.append(
            {
                "record_type": "sheet_name_mapping",
                "record_id": f"sheet_{index:02d}",
                "name": logical_name,
                "value": excel_name,
                "notes": "Explicit Excel-safe mapping (maximum 31 characters).",
            }
        )
    methodology_records.extend(_formula_records(formulas))
    methodology_records.extend(_methodology_records(methodology))
    methodology_records.extend(exact_machine_records)
    methodology_records.extend(rounding_records)
    methodology_records.extend(review_registry_records)
    methodology_records.extend(
        {
            "record_type": "artifact_manifest",
            "record_id": f"artifact_{index:04d}",
            "name": record["artifact_name"],
            "value": record["manifest_status"],
            "artifact_path": record["path"],
            "artifact_href": record["href"],
            "artifact_sha256": record["sha256"],
            "artifact_bytes": record["bytes"],
            "media_type": record["media_type"],
            "notes": (
                "Hash excluded by explicit manifest policy."
                if record["manifest_status"] == "excluded_by_policy"
                else "SHA-256 verified from payload"
                if record["sha256"]
                else "Hash unavailable; review required."
            ),
        }
        for index, record in enumerate(manifest, start=1)
    )
    range_records = [
        {
            "record_type": "table_range",
            "record_id": f"range_{index:02d}",
            "name": logical_name,
            "scope": SHEET_NAME_MAP[logical_name],
            "notes": f"Defined name: report_range_{index:02d}",
        }
        for index, logical_name in enumerate(SHEET_NAME_MAP, start=1)
    ]
    methodology_records.extend(range_records)
    methodology_rows = [[record.get(header) for header in _METHODOLOGY_HEADERS] for record in methodology_records]
    tables["Methodology"] = list(_METHODOLOGY_HEADERS), methodology_rows
    for logical_name, range_record in zip(SHEET_NAME_MAP, range_records, strict=True):
        headers, rows = tables[logical_name]
        sheet_range = _range_for(headers, rows)
        excel_name = SHEET_NAME_MAP[logical_name]
        range_record["range"] = f"'{excel_name}'!{sheet_range}"
    # Rebuild after the range placeholders were populated.  Its row count is
    # unchanged, so the Methodology self-range remains exact.
    tables["Methodology"] = (
        list(_METHODOLOGY_HEADERS),
        [[record.get(header) for header in _METHODOLOGY_HEADERS] for record in methodology_records],
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Soil Stamp Antonov"
    workbook.properties.lastModifiedBy = "Soil Stamp Antonov"
    workbook.properties.title = title or "Soil Stamp Antonov engineering report"
    workbook.properties.subject = "Deterministic plate-load-test report package"
    workbook.properties.created = _FIXED_TIMESTAMP
    workbook.properties.modified = _FIXED_TIMESTAMP
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"

    thin_gray = Side(style="thin", color="B7B7B7")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    review_fill = PatternFill("solid", fgColor="F4B183")
    review_font = Font(color="9C0006", bold=True)
    ready_fill = PatternFill("solid", fgColor="C6E0B4")
    table_style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    for sheet_index, logical_name in enumerate(SHEET_NAME_MAP, start=1):
        excel_name = SHEET_NAME_MAP[logical_name]
        headers, rows = tables[logical_name]
        worksheet = workbook.create_sheet(excel_name)
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = f"A{_DATA_ROW}"
        worksheet.auto_filter.ref = _range_for(headers, rows)
        column_count = max(1, len(headers))
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=column_count,
        )
        banner = worksheet.cell(row=1, column=1)
        banner.value = (
            "REVIEW REQUIRED — engineering approval is required before use."
            if effective_review
            else "NO REVIEW_REQUIRED FLAG — exported values are ready for documented review."
        )
        banner.data_type = "s"
        banner.fill = review_fill if effective_review else ready_fill
        banner.font = review_font if effective_review else Font(bold=True, color="006100")
        banner.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 24
        note = worksheet.cell(row=2, column=1)
        note.value = (
            "Raw data is a lossless report view of the values supplied to the renderer; "
            "the original source artifact remains authoritative."
            if logical_name == "Raw data"
            else "Machine numeric cells retain exported precision; display-only rounding is explicit."
        )
        note.data_type = "s"
        note.font = Font(italic=True, color="666666")

        effective_headers = headers or ["value"]
        for column_index, header in enumerate(effective_headers, start=1):
            cell = worksheet.cell(row=_HEADER_ROW, column=column_index)
            _set_literal(cell, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)

        for row_index, row in enumerate(rows, start=_DATA_ROW):
            for column_index, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                _set_literal(cell, value)
                if isinstance(cell.value, (int, float, Decimal)) and not isinstance(cell.value, bool):
                    cell.number_format = _MACHINE_NUMBER_FORMAT
                if isinstance(cell.value, str) and "review_required" in cell.value.casefold():
                    cell.fill = review_fill
                    cell.font = review_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if rows and headers:
            table = Table(
                displayName=f"ReportTable{sheet_index:02d}",
                ref=_range_for(headers, rows),
            )
            table.tableStyleInfo = table_style
            worksheet.add_table(table)

        for column_index, header in enumerate(effective_headers, start=1):
            values = [str(header)]
            values.extend(
                "" if column_index > len(row) or _is_missing(row[column_index - 1])
                else str(row[column_index - 1])
                for row in rows[:250]
            )
            width = min(60, max(10, max(len(value) for value in values) + 2))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        range_name = f"report_range_{sheet_index:02d}"
        sheet_range = _range_for(headers, rows)
        workbook.defined_names.add(
            DefinedName(range_name, attr_text=f"'{excel_name}'!{sheet_range}")
        )

    for sheet_name in (SHEET_NAME_MAP["Plots index"], SHEET_NAME_MAP["Methodology"]):
        worksheet = workbook[sheet_name]
        header_by_name = {
            str(worksheet.cell(_HEADER_ROW, column).value): column
            for column in range(1, worksheet.max_column + 1)
        }
        href_column = header_by_name.get("href") or header_by_name.get("artifact_href")
        if href_column is None:
            continue
        for row_index in range(_DATA_ROW, worksheet.max_row + 1):
            cell = worksheet.cell(row_index, href_column)
            if isinstance(cell.value, str) and cell.value:
                if _safe_package_hyperlink(cell.value):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    output = io.BytesIO()
    workbook.save(output)
    return _normalise_zip(output.getvalue())


__all__ = ["SHEET_NAME_MAP", "build_xlsx_report_package"]
