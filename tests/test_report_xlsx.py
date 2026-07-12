from __future__ import annotations

import hashlib
import zipfile
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from soilstamp.report_xlsx import SHEET_NAME_MAP, build_xlsx_report_package


def _open(payload: bytes):
    return load_workbook(BytesIO(payload), data_only=False)


def _sheet_records(worksheet) -> list[dict[str, object]]:
    headers = [
        worksheet.cell(row=3, column=column).value
        for column in range(1, worksheet.max_column + 1)
    ]
    return [
        {
            str(header): worksheet.cell(row=row, column=column).value
            for column, header in enumerate(headers, start=1)
            if header is not None
        }
        for row in range(4, worksheet.max_row + 1)
    ]


def _header_columns(worksheet) -> dict[str, int]:
    return {
        str(worksheet.cell(row=3, column=column).value): column
        for column in range(1, worksheet.max_column + 1)
    }


def _base_kwargs() -> dict[str, object]:
    raw = pd.DataFrame(
        {
            "test_id": ["0007", "0008", "0009", "0010", None],
            "reading": ["1,25", "=1+1", "+CMD", "-2+3", None],
            "comment": ["@SUM(A1:A2)", "literal", "", None, "safe"],
        },
        dtype=object,
    )
    prepared = pd.DataFrame(
        {
            "test_id": ["0007"],
            "pressure_kPa": [123.456789012345],
            "settlement_mm": [0.0123456789012345],
            "review_status": ["review_required"],
        }
    )
    return {
        "metadata": {
            "project": "XLSX test",
            "project_passport": {"project_id": "P-001", "operator": "=malicious"},
        },
        "raw": raw,
        "prepared": prepared,
        "indicator_passports": pd.DataFrame(
            [{"channel": "indicator_1", "serial_number": "0000123"}]
        ),
        "indicator_audit": pd.DataFrame(
            [{"test_id": "0007", "raw_reading": "1,25", "increment_mm": 0.01}]
        ),
        "qc_issues": [{"level": "warning", "code": "check", "message": "Review"}],
        "failures": pd.DataFrame(
            [{"test_id": "0007", "censoring_type": "right_censored", "lower_bound": 10.0}]
        ),
        "pcr_results": {
            "0007": {
                "pcr_auto": 51.2345678901234,
                "used_indices": [0, 1, 2],
            }
        },
        "moduli": pd.DataFrame(
            [{"test_id": "0007", "E_stamp_app_kPa": 12345.6789012345}]
        ),
        "group_comparisons": [
            pd.DataFrame([{"baseline_group": "B", "reinforced_group": "R", "n": 2}])
        ],
        "audit": [
            {
                "event_id": 1,
                "timestamp_utc": "2026-07-12T00:00:00+00:00",
                "action": "confirm",
            }
        ],
        "provenance": {
            "program_version": "test",
            "input_file_sha256": "b" * 64,
        },
        "formulas": {
            "pressure": {
                "expression": "=F_kN/A_m2",
                "scope": "Prepared data.pressure_kPa",
                "units": "kPa",
            }
        },
        "display_rounding": {"Prepared data": {"pressure_kPa": 2}},
    }


def test_xlsx_has_all_safe_sheets_and_lossless_raw_literals() -> None:
    workbook = _open(build_xlsx_report_package(**_base_kwargs()))

    assert workbook.sheetnames == list(SHEET_NAME_MAP.values())
    assert len(workbook.sheetnames) == 14
    assert all(len(name) <= 31 and "/" not in name for name in workbook.sheetnames)

    raw = workbook[SHEET_NAME_MAP["Raw data"]]
    assert [raw.cell(row=row, column=1).value for row in range(4, 9)] == [
        "0007",
        "0008",
        "0009",
        "0010",
        None,
    ]
    assert [raw.cell(row=row, column=2).value for row in range(4, 9)] == [
        "1,25",
        "=1+1",
        "+CMD",
        "-2+3",
        None,
    ]
    assert raw["A4"].data_type == "s"
    assert raw["B4"].data_type == "s"
    for coordinate in ("B5", "B6", "B7", "C4"):
        assert raw[coordinate].data_type == "s"
        assert raw[coordinate].quotePrefix is True

    formula_cells = [
        cell.coordinate
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    assert formula_cells == []


def test_machine_precision_display_rounding_status_formula_and_ranges_are_explicit() -> None:
    workbook = _open(build_xlsx_report_package(**_base_kwargs()))
    prepared = workbook[SHEET_NAME_MAP["Prepared data"]]

    assert prepared["B4"].value == 123.456789012345
    assert prepared["B4"].data_type == "n"
    assert prepared["B4"].number_format == "0.###############"
    columns = _header_columns(prepared)
    exact_cell = prepared.cell(4, columns["pressure_kPa__exact_machine"])
    display_cell = prepared.cell(4, columns["pressure_kPa__display"])
    assert exact_cell.value == "123.456789012345"
    assert exact_cell.data_type == "s"
    assert display_cell.value == "123,46"
    assert display_cell.data_type == "s"

    assert prepared["A1"].value.startswith("REVIEW REQUIRED")
    assert prepared["A1"].fill.fill_type == "solid"
    assert prepared["A1"].fill.fgColor.rgb.endswith("F4B183")
    assert prepared["D4"].fill.fgColor.rgb.endswith("F4B183")

    records = _sheet_records(workbook[SHEET_NAME_MAP["Methodology"]])
    mapping = {
        row["name"]: row["value"]
        for row in records
        if row["record_type"] == "sheet_name_mapping"
    }
    assert mapping == SHEET_NAME_MAP
    formula = next(row for row in records if row["record_type"] == "formula")
    assert formula["record_id"] == "pressure"
    assert formula["expression"] == "=F_kN/A_m2"
    assert formula["units"] == "kPa"

    ranges = [row for row in records if row["record_type"] == "table_range"]
    assert len(ranges) == len(SHEET_NAME_MAP)
    assert {row["name"] for row in ranges} == set(SHEET_NAME_MAP)
    assert all(str(row["range"]).startswith("'") and "!$A$3:" in row["range"] for row in ranges)
    assert list(workbook.defined_names) == [
        f"report_range_{index:02d}" for index in range(1, len(SHEET_NAME_MAP) + 1)
    ]


def test_artifact_manifest_hashes_every_artifact_and_plot_index_has_links() -> None:
    source_payload = b"exact source bytes\x00"
    plot_payload = b"<svg xmlns='http://www.w3.org/2000/svg'/>"
    kwargs = _base_kwargs()
    kwargs.update(
        {
            "artifacts": [
                {
                    "path": "source/protocol.csv",
                    "href": "source/protocol.csv",
                    "payload": source_payload,
                    "media_type": "text/csv",
                },
                {
                    "path": "artifact_manifest.json",
                    "href": "artifact_manifest.json",
                    "bytes": None,
                    "sha256": "excluded:self-referential-manifest",
                    "media_type": "application/json",
                },
            ],
            # Facade-compatible record: ``bytes`` is a count, not payload.
            "plots": [
                {
                    "path": "figures/antonov.svg",
                    "href": "figures/antonov.svg",
                    "bytes": len(plot_payload),
                    "sha256": hashlib.sha256(plot_payload).hexdigest(),
                    "media_type": "image/svg+xml",
                }
            ],
        }
    )
    workbook = _open(build_xlsx_report_package(**kwargs))

    methodology = _sheet_records(workbook[SHEET_NAME_MAP["Methodology"]])
    manifest = {
        row["artifact_path"]: row
        for row in methodology
        if row["record_type"] == "artifact_manifest"
    }
    assert set(manifest) == {
        "artifact_manifest.json",
        "source/protocol.csv",
        "figures/antonov.svg",
    }
    assert manifest["source/protocol.csv"]["artifact_sha256"] == hashlib.sha256(
        source_payload
    ).hexdigest()
    assert manifest["source/protocol.csv"]["artifact_bytes"] == len(source_payload)
    assert manifest["figures/antonov.svg"]["artifact_sha256"] == hashlib.sha256(
        plot_payload
    ).hexdigest()
    assert manifest["artifact_manifest.json"]["value"] == "excluded_by_policy"

    plots = workbook[SHEET_NAME_MAP["Plots index"]]
    plot_records = _sheet_records(plots)
    assert [record["path"] for record in plot_records] == ["figures/antonov.svg"]
    href_column = next(
        column for column in range(1, plots.max_column + 1) if plots.cell(3, column).value == "href"
    )
    assert plots.cell(4, href_column).hyperlink.target == "figures/antonov.svg"


def test_xlsx_bytes_and_ooxml_metadata_are_deterministic() -> None:
    kwargs = _base_kwargs()
    first = build_xlsx_report_package(**kwargs)
    second = build_xlsx_report_package(**kwargs)

    assert first == second
    workbook = _open(first)
    assert workbook.properties.created == datetime(2000, 1, 1)
    assert workbook.properties.modified == datetime(2000, 1, 1)
    with zipfile.ZipFile(BytesIO(first)) as archive:
        assert archive.namelist() == sorted(archive.namelist())
        assert {item.date_time for item in archive.infolist()} == {(2000, 1, 1, 0, 0, 0)}


def test_result_table_aliases_title_and_explicit_clear_status_are_supported() -> None:
    payload = build_xlsx_report_package(
        metadata={},
        raw=pd.DataFrame({"raw": ["01"]}),
        prepared=pd.DataFrame({"machine": [1.5]}),
        result_tables={
            "failure_summary": pd.DataFrame(
                [{"test_id": "T-1", "censoring_type": "interval_censored"}]
            )
        },
        title="Approved package view",
        review_required=False,
    )
    workbook = _open(payload)

    assert workbook.properties.title == "Approved package view"
    assert workbook["Project passport"]["A1"].value.startswith("NO REVIEW_REQUIRED FLAG")
    failures = _sheet_records(workbook[SHEET_NAME_MAP["Failure/censoring"]])
    assert failures == [{"test_id": "T-1", "censoring_type": "interval_censored"}]


def test_review_required_registry_is_preserved_with_scope_and_message() -> None:
    payload = build_xlsx_report_package(
        metadata={},
        raw=pd.DataFrame(),
        prepared=pd.DataFrame(),
        review_required=[
            {
                "category": "modulus_methodology",
                "scope": "T-07",
                "status": "review_required",
                "message": "Confirm the approved pressure range.",
            }
        ],
    )
    workbook = _open(payload)
    records = _sheet_records(workbook[SHEET_NAME_MAP["Methodology"]])
    review = next(row for row in records if row["record_type"] == "review_required")

    assert workbook["Project passport"]["A1"].value.startswith("REVIEW REQUIRED")
    assert review["name"] == "modulus_methodology"
    assert review["scope"] == "T-07"
    assert review["notes"] == "Confirm the approved pressure range."


def test_machine_exact_repr_survives_excel_precision_and_nonfinite_states_are_literal() -> None:
    exact_decimal = Decimal("1234567890.12345678901234567890")
    raw = pd.DataFrame(
        {"state": pd.Series([float("nan"), float("inf"), float("-inf")], dtype=object)}
    )
    prepared = pd.DataFrame(
        {
            "high_precision": pd.Series([exact_decimal, None, None], dtype=object),
            "state": pd.Series(
                [float("nan"), float("inf"), float("-inf")],
                dtype=object,
            ),
        }
    )
    workbook = _open(
        build_xlsx_report_package(metadata={}, raw=raw, prepared=prepared)
    )

    raw_sheet = workbook[SHEET_NAME_MAP["Raw data"]]
    assert [raw_sheet.cell(row, 1).value for row in range(4, 7)] == [
        "NaN",
        "+Inf",
        "-Inf",
    ]
    assert all(raw_sheet.cell(row, 1).data_type == "s" for row in range(4, 7))

    machine = workbook[SHEET_NAME_MAP["Prepared data"]]
    columns = _header_columns(machine)
    numeric = machine.cell(4, columns["high_precision"])
    exact = machine.cell(4, columns["high_precision__exact_machine"])
    assert numeric.data_type == "n"
    assert numeric.value != exact_decimal
    assert exact.value == str(exact_decimal)
    assert exact.data_type == "s"
    assert [machine.cell(row, columns["state"]).value for row in range(4, 7)] == [
        "NaN",
        "+Inf",
        "-Inf",
    ]
    assert [
        machine.cell(row, columns["state__exact_machine"]).value
        for row in range(4, 7)
    ] == ["NaN", "+Inf", "-Inf"]


def test_only_relative_percent_checked_package_paths_become_hyperlinks() -> None:
    unsafe_hrefs = [
        "https://example.invalid/plot.svg",
        "file:///C:/secret.svg",
        "javascript:alert(1)",
        "//server/share/plot.svg",
        "\\\\server\\share\\plot.svg",
        "/root/plot.svg",
        "../escape.svg",
        "%2e%2e/escape.svg",
        "%252e%252e/escape.svg",
        "C:/drive.svg",
    ]
    records = [
        {
            "name": "safe_plot",
            "path": "figures/safe plot.svg",
            "href": "figures/safe%20plot.svg",
            "bytes": 1,
            "sha256": "a" * 64,
            "media_type": "image/svg+xml",
        }
    ]
    records.extend(
        {
            "name": f"unsafe_{index:02d}",
            "path": f"figures/unsafe_{index:02d}.svg",
            "href": href,
            "bytes": 1,
            "sha256": f"{index + 1:064x}",
            "media_type": "image/svg+xml",
        }
        for index, href in enumerate(unsafe_hrefs)
    )
    workbook = _open(
        build_xlsx_report_package(
            metadata={},
            raw=pd.DataFrame(),
            prepared=pd.DataFrame(),
            plots=records,
        )
    )
    plots = workbook[SHEET_NAME_MAP["Plots index"]]
    columns = _header_columns(plots)
    rows_by_name = {
        plots.cell(row, columns["artifact_name"]).value: row
        for row in range(4, plots.max_row + 1)
    }

    safe_cell = plots.cell(rows_by_name["safe_plot"], columns["href"])
    assert safe_cell.value == "figures/safe%20plot.svg"
    assert safe_cell.hyperlink.target == "figures/safe%20plot.svg"
    for index, href in enumerate(unsafe_hrefs):
        cell = plots.cell(rows_by_name[f"unsafe_{index:02d}"], columns["href"])
        assert cell.value == href
        assert cell.hyperlink is None
