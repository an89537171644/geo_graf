from __future__ import annotations

import hashlib
import zipfile
from io import BytesIO
from pathlib import Path

import numpy as np
import pytest
from openpyxl import Workbook

from soilstamp.data import prepare_measurements as _prepare_measurements
from soilstamp.io import (
    read_metadata_json,
    read_protocol,
    read_protocol_excel,
    validate_import_metadata_consistency,
)


def prepare_measurements(*args, **kwargs):
    kwargs.setdefault("strict_metadata", False)
    return _prepare_measurements(*args, **kwargs)


def _workbook_bytes(rows: list[list[object]], title: str = "Протокол") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_metadata_json_rejects_duplicate_keys_at_any_depth() -> None:
    with pytest.raises(ValueError, match="повторяющийся ключ 'load_unit'"):
        read_metadata_json(b'{"load_unit":"kN","load_unit":"N"}')
    with pytest.raises(ValueError, match="повторяющийся ключ 'group'"):
        read_metadata_json(b'{"tests":{"T1":{"group":"a","group":"b"}}}')


def test_strict_excel_russian_headers_decimal_comma_and_source_cells() -> None:
    payload = _workbook_bytes(
        [
            ["Осадка, мм", "№ испытания", "Нагрузка, кН", "Ступень", "Статус"],
            ["0,25", "T-01", "1,5", 1, "stable"],
            ["0,40", "T-01", "2,0", 1, "stable"],
            ["0,60", "T-01", "2,5", None, "stable"],
        ]
    )
    result = read_protocol_excel(payload, import_mode="strict")

    assert not result.blocking_issues
    assert list(result.frame["test_id"]) == ["T-01", "T-01", "T-01"]
    assert np.allclose(result.frame["load"], [1.5, 2.0, 2.5])
    assert np.allclose(result.frame["settlement"], [0.25, 0.40, 0.60])
    assert list(result.frame["stage"][:2]) == [1, 1]
    assert result.frame.loc[0, "raw_load"] == "1,5"
    assert result.frame.loc[0, "parsed_load"] == 1.5
    assert result.frame.loc[0, "source_row"] == 2
    assert result.frame.loc[0, "load_unit"] == "kN"
    assert len(result.raw_cells) == 15
    assert any(issue.code == "missing_stage" and issue.row == 4 for issue in result.issues)
    _, measurement_issues = prepare_measurements(
        result.frame,
        {"load_kind": "force", "load_unit": "kN", "stamp_area_m2": 0.1},
    )
    assert any(issue.code == "repeated_stage" for issue in measurement_issues)


def test_strict_excel_preserves_explicit_indicator_turn_number() -> None:
    payload = _workbook_bytes(
        [
            [
                "test_id",
                "stage",
                "load",
                "indicator_1",
                "номер оборота индикатора 1",
            ],
            ["T1", 0, 0.0, 0.0, 0],
            ["T1", 1, 1.0, 0.2, 2],
        ]
    )

    result = read_protocol_excel(payload, import_mode="strict")

    assert not result.blocking_issues
    assert result.frame["indicator_1_turn_number"].tolist() == [0.0, 2.0]


def test_strict_unknown_header_blocks_with_exact_location() -> None:
    payload = _workbook_bytes(
        [
            ["test_id", "stage", "load", "settlement", "Неизвестное поле"],
            ["T1", 1, 1.0, 0.1, "x"],
        ]
    )
    result = read_protocol_excel(payload, import_mode="strict")
    issue = next(item for item in result.issues if item.code == "unknown_header")

    assert issue.blocks_processing is True
    assert issue.sheet == "Протокол"
    assert issue.row == 1
    assert issue.column == "E"
    assert issue.raw_value == "Неизвестное поле"


def test_interactive_excel_uses_saved_mapping() -> None:
    payload = _workbook_bytes(
        [
            ["Испытание", "Ступень", "Сила прибора", "Вертикальное перемещение"],
            ["A-7", "1", "1,25", "0,15"],
        ]
    )
    saved_mapping = {"test_id": "A", "stage": "B", "load": "C", "settlement": "D"}
    result = read_protocol_excel(
        payload,
        import_mode="interactive",
        column_mapping=saved_mapping,
    )

    assert not result.blocking_issues
    assert result.info["sheets"][0]["mapping"] == saved_mapping
    assert result.frame.loc[0, "test_id"] == "A-7"
    assert result.frame.loc[0, "load"] == 1.25
    assert result.frame.loc[0, "settlement"] == 0.15


def test_interactive_mapping_accepts_completely_unknown_headers() -> None:
    payload = _workbook_bytes(
        [["alpha", "beta", "gamma", "delta"], ["T9", 1, "2,5", "0,4"]]
    )
    result = read_protocol_excel(
        payload,
        import_mode="interactive",
        column_mapping={"test_id": "A", "stage": "B", "load": "C", "settlement": "D"},
        header_row=1,
    )

    assert not result.blocking_issues
    assert result.frame.loc[0, "test_id"] == "T9"
    assert result.frame.loc[0, "load"] == 2.5


def test_invalid_load_reports_exact_excel_cell() -> None:
    payload = _workbook_bytes(
        [["test_id", "stage", "load", "settlement"], ["T1", 1, "abc", 0.2]]
    )
    result = read_protocol_excel(payload, import_mode="strict")
    issue = next(item for item in result.issues if item.code == "invalid_load_cell")

    assert issue.sheet == "Протокол"
    assert issue.row == 2
    assert issue.column == "C"
    assert issue.raw_value == "abc"
    assert issue.blocks_processing is True


def test_invalid_settlement_text_is_blocking_with_exact_cell() -> None:
    payload = _workbook_bytes(
        [["test_id", "stage", "load", "settlement"], ["T1", 1, 1.0, "abc"]]
    )
    result = read_protocol_excel(payload, import_mode="strict")
    issue = next(item for item in result.issues if item.code == "invalid_measurement_cell")

    assert issue.blocks_processing is True
    assert issue.row == 2
    assert issue.column == "D"
    assert issue.raw_value == "abc"


def test_failure_text_in_settlement_survives_without_status_column() -> None:
    payload = _workbook_bytes(
        [
            ["test_id", "stage", "load", "settlement"],
            ["T1", 1, 1.0, 0.2],
            ["T1", 2, 2.0, "failure of specimen"],
        ]
    )
    imported = read_protocol_excel(payload, import_mode="strict")
    prepared, issues = prepare_measurements(
        imported.frame,
        {"load_kind": "force", "load_unit": "kN", "stamp_area_m2": 0.1},
    )

    assert not [item for item in issues if item.blocks_processing]
    assert prepared["is_failure"].tolist() == [False, True]
    assert np.isnan(prepared.loc[1, "settlement_mm"])
    assert imported.frame.loc[1, "status"] == "failure"
    assert imported.frame.loc[1, "failure_marker_raw"] == "failure of specimen"


def test_numeric_cells_reject_units_from_wrong_physical_family() -> None:
    payload = _workbook_bytes(
        [
            ["test_id", "stage", "load", "settlement"],
            ["T1", 1, "1mm", "0.1kN"],
        ]
    )
    result = read_protocol_excel(payload, import_mode="strict")

    assert any(item.code == "invalid_load_cell" for item in result.blocking_issues)
    assert any(item.code == "invalid_measurement_cell" for item in result.blocking_issues)


def test_numeric_cells_do_not_silently_strip_same_family_units() -> None:
    payload = _workbook_bytes(
        [
            ["test_id", "stage", "load", "settlement"],
            ["T1", 1, "1000N", "0.1cm"],
        ]
    )
    result = read_protocol_excel(payload, import_mode="strict")

    assert any(item.code == "invalid_load_cell" for item in result.blocking_issues)
    assert any(item.code == "invalid_measurement_cell" for item in result.blocking_issues)


def test_formula_raw_value_is_preserved_and_missing_cache_blocks() -> None:
    payload = _workbook_bytes(
        [["test_id", "stage", "load", "settlement"], ["T1", 1, "=1+1", 0.2]]
    )
    result = read_protocol_excel(payload, import_mode="strict")
    formula_cell = result.raw_cells[
        (result.raw_cells["source_row"] == 2)
        & (result.raw_cells["canonical_field"] == "load")
    ].iloc[0]

    assert formula_cell["raw_value"] == "=1+1"
    issue = next(item for item in result.blocking_issues if item.code == "formula_without_cached_value")
    assert issue.column == "C"


def test_legacy_formula_raw_value_is_preserved() -> None:
    payload = _workbook_bytes(
        [
            ["1 испытание"],
            ["Ступень", "Индикатор", "Нагрузка, кН"],
            [1, 9.8, "=1+1"],
        ]
    )
    result = read_protocol_excel(payload, import_mode="heuristic")
    load_cell = result.raw_cells[
        (result.raw_cells["source_row"] == 3)
        & (result.raw_cells["canonical_field"] == "load")
    ].iloc[0]

    assert load_cell["raw_value"] == "=1+1"
    assert any(item.code == "formula_without_cached_value" for item in result.blocking_issues)


def test_unknown_unit_and_missing_geometry_are_blocking_after_import() -> None:
    payload = _workbook_bytes(
        [["test_id", "stage", "load", "settlement"], ["T1", 1, 1.0, 0.2]]
    )
    imported = read_protocol_excel(payload, import_mode="strict")
    _, unknown_unit = prepare_measurements(
        imported.frame,
        {"load_kind": "force", "load_unit": "mystery", "stamp_diameter_mm": 300},
    )
    _, missing_geometry = prepare_measurements(
        imported.frame,
        {"load_kind": "force", "load_unit": "kN"},
    )

    assert any(item.code == "unsupported_load_unit" for item in unknown_unit)
    geometry_issue = next(item for item in missing_geometry if item.code == "missing_stamp_area")
    assert geometry_issue.level == "warning"


def test_excel_header_unit_conflict_with_metadata_is_blocking() -> None:
    payload = _workbook_bytes(
        [["test_id", "stage", "Нагрузка, кН", "settlement"], ["T1", 1, 1000, 0.2]]
    )
    imported = read_protocol_excel(payload, import_mode="strict")
    issues = validate_import_metadata_consistency(
        imported.frame,
        {"load_kind": "force", "load_unit": "N"},
        imported.info,
    )
    issue = next(item for item in issues if item.code == "load_unit_conflict")

    assert issue.blocks_processing is True
    assert issue.row == 1
    assert issue.column == "C"

    override_issues = validate_import_metadata_consistency(
        imported.frame,
        {
            "load_kind": "force",
            "load_unit": "kN",
            "tests": {"T1": {"load_unit": "N"}},
        },
        imported.info,
    )
    assert any(item.code == "load_unit_conflict" for item in override_issues)

    pressure_payload = _workbook_bytes(
        [["test_id", "stage", "load kPa", "settlement"], ["P1", 1, 25.0, 0.1]]
    )
    pressure_import = read_protocol_excel(pressure_payload, import_mode="strict")
    pressure_issues = validate_import_metadata_consistency(
        pressure_import.frame,
        {"load_kind": "pressure", "load_unit": "kPa"},
        pressure_import.info,
    )
    assert not pressure_import.blocking_issues
    assert not pressure_issues


def test_protocol_test_ids_must_match_metadata_passports_in_strict_mode() -> None:
    payload = _workbook_bytes(
        [["test_id", "stage", "load", "settlement"], ["T1", 1, 1.0, 0.2]]
    )
    imported = read_protocol_excel(payload, import_mode="strict")
    issues = validate_import_metadata_consistency(
        imported.frame,
        {"load_unit": "kN", "tests": {"OTHER": {"group": "baseline"}}},
        imported.info,
        strict=True,
    )

    assert any(
        item.code == "protocol_tests_missing_in_metadata" and item.blocks_processing
        for item in issues
    )


def test_excel_source_file_is_not_modified(tmp_path) -> None:
    payload = _workbook_bytes(
        [["test_id", "stage", "load", "settlement"], ["T1", 1, 1.0, 0.2]]
    )
    path = tmp_path / "исходный журнал.xlsx"
    path.write_bytes(payload)
    before = hashlib.sha256(path.read_bytes()).hexdigest()

    result = read_protocol_excel(path, import_mode="strict")
    after = hashlib.sha256(path.read_bytes()).hexdigest()

    assert not result.blocking_issues
    assert before == after == result.info["input_file_sha256"]


def test_heuristic_imports_legacy_test_blocks_without_fabricating_zero() -> None:
    payload = _workbook_bytes(
        [
            ["1 испытание"],
            ["Ступень", "Индикатор", "Нагрузка, кН"],
            [1, "9,80", "1,0"],
            [2, "9,55", "2,0"],
            [3, "ушла", "2,5"],
        ]
    )
    result = read_protocol_excel(payload, import_mode="heuristic")

    assert not result.blocking_issues
    assert list(result.frame["test_id"]) == ["1", "1", "1"]
    assert result.frame["status"].iloc[:2].isna().all()
    assert result.frame["status"].iloc[2] == "failure"
    assert not ((result.frame["load"] == 0) & (result.frame["indicator_1"] == 0)).any()
    assert any(item.code == "legacy_indicator_calibration_required" for item in result.issues)
    _, processing_issues = prepare_measurements(
        result.frame,
        {
            "load_kind": "force",
            "load_unit": "kN",
            "stamp_diameter_mm": 300,
        },
    )
    assert any(item.code == "uncalibrated_legacy_indicator" for item in processing_issues)


def test_legacy_failure_negation_is_not_classified_as_failure() -> None:
    payload = _workbook_bytes(
        [
            ["1 испытание"],
            ["Ступень", "Индикатор", "Нагрузка, кН"],
            [1, "не ушла", "1,0"],
        ]
    )
    result = read_protocol_excel(payload, import_mode="heuristic")

    assert result.frame.loc[0, "status"] is None


def test_duplicate_legacy_ids_are_not_silently_renamed() -> None:
    payload = _workbook_bytes(
        [
            ["1 испытание"],
            ["Ступень", "Индикатор", "Нагрузка, кН"],
            [1, 9.8, 1.0],
            ["1 испытание"],
            ["Ступень", "Индикатор", "Нагрузка, кН"],
            [1, 9.7, 1.0],
        ]
    )
    result = read_protocol_excel(payload, import_mode="heuristic")

    assert set(result.frame["test_id"]) == {"1"}
    assert set(result.frame["test_id_raw"]) == {"1"}
    assert any(item.code == "duplicate_legacy_test_id" for item in result.blocking_issues)


def test_strict_workbook_skips_service_sheet_when_protocol_sheet_is_valid() -> None:
    workbook = Workbook()
    protocol = workbook.active
    protocol.title = "Protocol"
    protocol.append(["test_id", "stage", "load", "settlement"])
    protocol.append(["T1", 1, 1.0, 0.2])
    notes = workbook.create_sheet("Notes")
    notes.append(["Служебная информация"])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    result = read_protocol_excel(buffer.getvalue(), import_mode="strict")

    assert not result.blocking_issues
    assert len(result.frame) == 1
    assert any(item.code == "sheet_skipped_no_schema" for item in result.issues)


def test_duplicate_test_id_across_protocol_sheets_is_blocking() -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Run1"
    second = workbook.create_sheet("Run2")
    for sheet in (first, second):
        sheet.append(["test_id", "stage", "load", "settlement"])
        sheet.append(["T1", 1, 1.0, 0.2])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    result = read_protocol_excel(buffer.getvalue(), import_mode="strict")

    assert any(
        item.code == "duplicate_test_id_across_sheets" for item in result.blocking_issues
    )


def test_strict_does_not_silently_skip_unknown_tabular_sheet() -> None:
    workbook = Workbook()
    protocol = workbook.active
    protocol.title = "Protocol"
    protocol.append(["test_id", "stage", "load", "settlement"])
    protocol.append(["T1", 1, 1.0, 0.2])
    unknown = workbook.create_sheet("CustomData")
    unknown.append(["alpha", "beta"])
    unknown.append([1, 2])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    result = read_protocol_excel(buffer.getvalue(), import_mode="strict")

    assert any(item.code == "schema_not_recognized" for item in result.blocking_issues)


def test_handoff_legacy_demo_imports_in_compatibility_mode() -> None:
    fixture = Path(__file__).parent / "fixtures" / "legacy_demo_input.xlsx"
    result = read_protocol_excel(fixture, import_mode="heuristic")

    assert not result.blocking_issues
    assert len(result.frame) == 69
    assert result.frame["test_id"].nunique() == 6
    assert any(item.code == "heuristic_import" for item in result.issues)
    assert result.frame["indicator_requires_calibration"].all()


def test_corrupt_xlsx_container_returns_controlled_blocking_issue() -> None:
    result = read_protocol_excel(b"PK\x03\x04not-a-valid-archive", import_mode="strict")

    assert result.frame.empty
    assert [item.code for item in result.blocking_issues] == ["invalid_xlsx_container"]


def test_xlsx_zip_bomb_ratio_is_rejected_before_openpyxl() -> None:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
        archive.writestr("xl/workbook.xml", "<workbook/>")
        archive.writestr("xl/sharedStrings.xml", b"0" * (2 * 1024 * 1024))

    result = read_protocol_excel(buffer.getvalue(), import_mode="strict")

    assert result.frame.empty
    assert any(item.code == "xlsx_resource_limit" for item in result.blocking_issues)


def test_xlsx_xml_entity_is_rejected_by_safe_parser() -> None:
    source = _workbook_bytes(
        [["test_id", "stage", "load", "settlement"], ["T1", 1, 1.0, 0.2]]
    )
    source_archive = zipfile.ZipFile(BytesIO(source))
    buffer = BytesIO()
    with source_archive, zipfile.ZipFile(
        buffer, "w", compression=zipfile.ZIP_DEFLATED
    ) as target:
        for item in source_archive.infolist():
            payload = source_archive.read(item.filename)
            if item.filename == "xl/workbook.xml":
                declaration_end = payload.find(b"?>")
                insertion = declaration_end + 2 if declaration_end >= 0 else 0
                payload = (
                    payload[:insertion]
                    + b'<!DOCTYPE workbook [<!ENTITY unsafe "blocked">]>'
                    + payload[insertion:]
                )
            target.writestr(item.filename, payload)

    result = read_protocol_excel(buffer.getvalue(), import_mode="strict")

    assert result.frame.empty
    assert any(item.code == "unsafe_xml" for item in result.blocking_issues)


def test_declared_worksheet_dimensions_are_limited() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=200_001, column=1, value="x")
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()

    result = read_protocol_excel(buffer.getvalue(), import_mode="strict")

    assert result.frame.empty
    assert any(item.code == "xlsx_resource_limit" for item in result.blocking_issues)


def test_xlsm_is_read_without_executing_macros_and_warns() -> None:
    payload = _workbook_bytes(
        [["test_id", "stage", "load", "settlement"], ["T1", 1, 1.0, 0.2]]
    )
    result = read_protocol(payload, filename="protocol.xlsm", import_mode="strict")

    assert len(result.frame) == 1
    assert any(item.code == "xlsm_macros_ignored" for item in result.issues)
    assert result.info["macros_ignored"] is True
